from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy import func
from sqlmodel import Session, select

from app.enums import AcceptanceResult, ClaimStatus, ProblemStatus, RewardRoleType, RewardStatus
from app.models import (
    Acceptance,
    Claim,
    Knowledge,
    Problem,
    Reward,
    Task,
    User,
)
from app.schemas import (
    DashboardDistribution,
    DashboardRankings,
    DashboardTrends,
    DistributionItem,
    RankingItem,
    TimeRange,
    TrendGranularity,
    TrendPoint,
)


def _start_of_time_range(time_range: TimeRange) -> datetime | None:
    now = datetime.utcnow()
    if time_range == "all":
        return None
    if time_range == "this_year":
        return datetime(now.year, 1, 1)
    if time_range == "this_month":
        return datetime(now.year, now.month, 1)
    quarter_start_month = ((now.month - 1) // 3) * 3 + 1
    return datetime(now.year, quarter_start_month, 1)


def _append_time_filter(statement, column, time_range: TimeRange):
    start = _start_of_time_range(time_range)
    if start is not None:
        statement = statement.where(column >= start)
    return statement


def _to_ranking(rows: list[tuple[int, str, float]]) -> list[RankingItem]:
    return [
        RankingItem(user_id=int(row[0]), user_name=row[1], value=float(row[2]))
        for row in rows
    ]


def dashboard_rankings(session: Session, time_range: TimeRange, top_n: int = 10) -> DashboardRankings:
    claim_stmt = (
        select(Claim.lead_user_id, User.name, func.count(Claim.id).label("value"))
        .join(User, User.id == Claim.lead_user_id)
        .where(Claim.status == ClaimStatus.COMPLETED)
    )
    claim_stmt = _append_time_filter(claim_stmt, Claim.created_at, time_range)
    claim_rows = session.exec(
        claim_stmt.group_by(Claim.lead_user_id, User.name).order_by(func.count(Claim.id).desc()).limit(top_n)
    ).all()

    reward_stmt = (
        select(Reward.user_id, User.name, func.coalesce(func.sum(Reward.amount), 0).label("value"))
        .join(User, User.id == Reward.user_id)
        .where(Reward.status == RewardStatus.CONFIRMED)
    )
    reward_stmt = _append_time_filter(reward_stmt, Reward.confirmed_at, time_range)
    reward_rows = session.exec(
        reward_stmt.group_by(Reward.user_id, User.name)
        .order_by(func.sum(Reward.amount).desc())
        .limit(top_n)
    ).all()

    problem_stmt = (
        select(Problem.submitter_id, User.name, func.count(Problem.id).label("value"))
        .join(User, User.id == Problem.submitter_id)
        .where(Problem.status.in_([ProblemStatus.APPROVED, ProblemStatus.ARCHIVED]))
    )
    problem_stmt = _append_time_filter(problem_stmt, Problem.created_at, time_range)
    problem_rows = session.exec(
        problem_stmt.group_by(Problem.submitter_id, User.name)
        .order_by(func.count(Problem.id).desc())
        .limit(top_n)
    ).all()

    points_stmt = (
        select(Reward.user_id, User.name, func.coalesce(func.sum(Reward.points), 0).label("value"))
        .join(User, User.id == Reward.user_id)
        .where(Reward.status == RewardStatus.CONFIRMED)
    )
    points_stmt = _append_time_filter(points_stmt, Reward.confirmed_at, time_range)
    points_rows = session.exec(
        points_stmt.group_by(Reward.user_id, User.name)
        .order_by(func.sum(Reward.points).desc())
        .limit(top_n)
    ).all()

    return DashboardRankings(
        claim_count_ranking=_to_ranking(claim_rows),
        reward_amount_ranking=_to_ranking(reward_rows),
        problem_contribution_ranking=_to_ranking(problem_rows),
        points_ranking=_to_ranking(points_rows),
    )


def dashboard_trends(
    session: Session, time_range: TimeRange, granularity: TrendGranularity
) -> DashboardTrends:
    pattern = "%Y-%m" if granularity == "month" else "%Y-W%W"
    period_problem = func.strftime(pattern, Problem.created_at)
    period_acceptance = func.strftime(pattern, Acceptance.created_at)
    period_reward = func.strftime(pattern, Reward.confirmed_at)

    problem_stmt = select(period_problem, func.count(Problem.id)).group_by(period_problem)
    problem_stmt = _append_time_filter(problem_stmt, Problem.created_at, time_range)
    problem_rows = session.exec(problem_stmt).all()

    completed_stmt = (
        select(period_acceptance, func.count(Acceptance.id))
        .where(Acceptance.result == AcceptanceResult.APPROVED.value)
        .group_by(period_acceptance)
    )
    completed_stmt = _append_time_filter(completed_stmt, Acceptance.created_at, time_range)
    completed_rows = session.exec(completed_stmt).all()

    reward_stmt = (
        select(period_reward, func.coalesce(func.sum(Reward.amount), 0))
        .where(Reward.status == RewardStatus.CONFIRMED)
        .group_by(period_reward)
    )
    reward_stmt = _append_time_filter(reward_stmt, Reward.confirmed_at, time_range)
    reward_rows = session.exec(reward_stmt).all()

    problem_map = {row[0]: int(row[1]) for row in problem_rows if row[0]}
    completed_map = {row[0]: int(row[1]) for row in completed_rows if row[0]}
    reward_map = {row[0]: float(row[1]) for row in reward_rows if row[0]}

    periods = sorted(set(problem_map.keys()) | set(completed_map.keys()) | set(reward_map.keys()))
    points = [
        TrendPoint(
            period=period,
            problem_submitted=problem_map.get(period, 0),
            task_completed=completed_map.get(period, 0),
            reward_confirmed_amount=reward_map.get(period, 0.0),
        )
        for period in periods
    ]
    return DashboardTrends(granularity=granularity, points=points)


def dashboard_distribution(session: Session, time_range: TimeRange) -> DashboardDistribution:
    scenario_stmt = select(Problem.scenario, func.count(Problem.id)).group_by(Problem.scenario)
    scenario_stmt = _append_time_filter(scenario_stmt, Problem.created_at, time_range)
    scenario_rows = session.exec(scenario_stmt).all()

    level_stmt = select(Task.level, func.count(Task.id)).group_by(Task.level)
    level_stmt = _append_time_filter(level_stmt, Task.created_at, time_range)
    level_rows = session.exec(level_stmt).all()

    dept_stmt = (
        select(func.coalesce(User.department, "UNKNOWN"), func.count(Claim.id))
        .join(User, User.id == Claim.lead_user_id)
        .where(Claim.status == ClaimStatus.COMPLETED)
        .group_by(User.department)
    )
    dept_stmt = _append_time_filter(dept_stmt, Claim.created_at, time_range)
    dept_rows = session.exec(dept_stmt).all()

    return DashboardDistribution(
        scenario_distribution=[
            DistributionItem(name=str(row[0]), count=int(row[1])) for row in scenario_rows
        ],
        level_distribution=[DistributionItem(name=str(row[0]), count=int(row[1])) for row in level_rows],
        department_distribution=[DistributionItem(name=str(row[0]), count=int(row[1])) for row in dept_rows],
    )


def _workbook_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _append_table(sheet, headers: list[str], rows: list[list]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def export_tasks_excel(session: Session) -> bytes:
    rows = session.exec(select(Task).order_by(Task.created_at.desc())).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    _append_table(
        ws,
        ["task_id", "problem_id", "title", "level", "reward_total", "due_date", "status", "created_at"],
        [
            [
                item.id,
                item.problem_id,
                item.title,
                item.level.value,
                item.reward_total,
                item.due_date.isoformat(),
                item.status.value,
                item.created_at.isoformat(),
            ]
            for item in rows
        ],
    )
    return _workbook_bytes(wb)


def export_rewards_excel(session: Session) -> bytes:
    rows = session.exec(select(Reward, User.name).join(User, User.id == Reward.user_id)).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Rewards"
    _append_table(
        ws,
        ["reward_id", "task_id", "user_id", "user_name", "role_type", "amount", "points", "badge", "status", "confirmed_at"],
        [
            [
                reward.id,
                reward.task_id,
                reward.user_id,
                user_name,
                reward.role_type.value,
                reward.amount,
                reward.points,
                reward.badge or "",
                reward.status.value,
                reward.confirmed_at.isoformat() if reward.confirmed_at else "",
            ]
            for reward, user_name in rows
        ],
    )
    return _workbook_bytes(wb)


def export_knowledge_excel(session: Session) -> bytes:
    rows = session.exec(select(Knowledge).order_by(Knowledge.archived_at.desc())).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Knowledge"
    _append_table(
        ws,
        ["knowledge_id", "task_id", "problem_summary", "solution_summary", "tags", "recommended", "archived_at"],
        [
            [
                item.id,
                item.task_id,
                item.problem_summary,
                item.solution_summary,
                item.tags,
                item.recommended,
                item.archived_at.isoformat(),
            ]
            for item in rows
        ],
    )
    return _workbook_bytes(wb)


def export_dashboard_excel(
    session: Session, time_range: TimeRange, granularity: TrendGranularity, top_n: int
) -> bytes:
    rankings = dashboard_rankings(session, time_range=time_range, top_n=top_n)
    trends = dashboard_trends(session, time_range=time_range, granularity=granularity)
    distribution = dashboard_distribution(session, time_range=time_range)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Rankings"
    _append_table(
        ws1,
        ["ranking_type", "user_id", "user_name", "value"],
        [
            ["claim_count", i.user_id, i.user_name, i.value]
            for i in rankings.claim_count_ranking
        ]
        + [
            ["reward_amount", i.user_id, i.user_name, i.value]
            for i in rankings.reward_amount_ranking
        ]
        + [
            ["problem_contribution", i.user_id, i.user_name, i.value]
            for i in rankings.problem_contribution_ranking
        ]
        + [["points", i.user_id, i.user_name, i.value] for i in rankings.points_ranking],
    )

    ws2 = wb.create_sheet("Trends")
    _append_table(
        ws2,
        ["period", "problem_submitted", "task_completed", "reward_confirmed_amount"],
        [
            [point.period, point.problem_submitted, point.task_completed, point.reward_confirmed_amount]
            for point in trends.points
        ],
    )

    ws3 = wb.create_sheet("Distribution")
    _append_table(
        ws3,
        ["distribution_type", "name", "count"],
        [["scenario", i.name, i.count] for i in distribution.scenario_distribution]
        + [["level", i.name, i.count] for i in distribution.level_distribution]
        + [["department", i.name, i.count] for i in distribution.department_distribution],
    )
    return _workbook_bytes(wb)


def export_knowledge_pdf(session: Session) -> bytes:
    rows = session.exec(select(Knowledge).order_by(Knowledge.archived_at.desc())).all()
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    width, height = A4
    y = height - 40
    pdf.setFont("Helvetica", 12)
    pdf.drawString(40, y, "Knowledge Export")
    y -= 24
    pdf.setFont("Helvetica", 10)
    for item in rows:
        line = f"#{item.id} task={item.task_id} recommended={item.recommended}"
        pdf.drawString(40, y, line[:110])
        y -= 16
        summary = f"P: {item.problem_summary} | S: {item.solution_summary}"
        pdf.drawString(40, y, summary[:110])
        y -= 20
        if y < 80:
            pdf.showPage()
            y = height - 40
            pdf.setFont("Helvetica", 10)
    pdf.save()
    return output.getvalue()
