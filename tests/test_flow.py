import os
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session, SQLModel, create_engine

from app.db import get_session
from app.auth import create_access_token
from app.enums import Role
from app.jobs import run_stale_progress_reminders
from app.main import app
from app.models import Claim, User, UserRole


def _headers(user_id: int) -> dict[str, str]:
    return {"X-User-Id": str(user_id)}


def _setup_client(tmp_path: Path) -> TestClient:
    db_file = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_file}", connect_args={"check_same_thread": False})
    SQLModel.metadata.create_all(engine)

    with Session(engine) as session:
        admin = User(id=1, name="Admin", employee_no="A001", department="PMO")
        session.add(admin)
        for role in [Role.ADMIN, Role.REVIEWER, Role.ACCEPTOR, Role.EMPLOYEE, Role.REWARD_APPROVER]:
            session.add(UserRole(user_id=1, role=role))
        session.commit()

    def override_session():
        with Session(engine) as session:
            yield session

    app.dependency_overrides[get_session] = override_session
    return TestClient(app)


def _review_problem_and_create_task(
    client: TestClient,
    reviewer_id: int,
    problem_id: int,
    payload: dict,
    budget_approver_id: int = 1,
    budget_comment: str = "budget approved",
):
    review_resp = client.post(
        f"/problems/{problem_id}/review",
        headers=_headers(reviewer_id),
        json=payload,
    )
    assert review_resp.status_code == 200
    assert review_resp.json()["status"] == "budget_pending"

    budget_approve_resp = client.post(
        f"/problems/{problem_id}/budget-review",
        headers=_headers(budget_approver_id),
        json={"approve": True, "comment": budget_comment},
    )
    assert budget_approve_resp.status_code == 200
    assert budget_approve_resp.json()["status"] == "approved"
    return budget_approve_resp


def _bearer_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def test_system_version_endpoint(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    resp = client.get("/system/version")
    assert resp.status_code == 200
    payload = resp.json()
    assert isinstance(payload["backend_version"], str)
    assert payload["backend_version"]
    assert isinstance(payload["backend_git_sha"], str)
    assert payload["backend_git_sha"]

    app.dependency_overrides.clear()


def test_auth_login_and_bearer_access(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    login_resp = client.post("/auth/login", json={"user_id": 1})
    assert login_resp.status_code == 200
    payload = login_resp.json()
    assert payload["token_type"] == "Bearer"
    assert payload["expires_in"] > 0
    assert payload["user"]["id"] == 1
    token = payload["access_token"]
    assert token

    me_resp = client.get("/me", headers=_bearer_headers(token))
    assert me_resp.status_code == 200
    assert me_resp.json()["id"] == 1

    app.dependency_overrides.clear()


def test_prod_auth_disables_passwordless_and_header_auth(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)
    os.environ["APP_ENV"] = "production"
    os.environ["AUTH_TOKEN_SECRET"] = "test-secret-for-production-mode"
    os.environ.pop("AUTH_ENABLE_PASSWORDLESS_LOGIN", None)
    os.environ.pop("AUTH_ENABLE_HEADER_USER_ID", None)
    try:
        login_resp = client.post("/auth/login", json={"user_id": 1})
        assert login_resp.status_code == 403

        me_header_resp = client.get("/me", headers=_headers(1))
        assert me_header_resp.status_code == 401

        token, _ = create_access_token(1)
        me_bearer_resp = client.get("/me", headers=_bearer_headers(token))
        assert me_bearer_resp.status_code == 200
        assert me_bearer_resp.json()["id"] == 1
    finally:
        os.environ.pop("APP_ENV", None)
        os.environ.pop("AUTH_TOKEN_SECRET", None)
        os.environ.pop("AUTH_ENABLE_PASSWORDLESS_LOGIN", None)
        os.environ.pop("AUTH_ENABLE_HEADER_USER_ID", None)
        app.dependency_overrides.clear()


def test_end_to_end_flow(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "Reviewer",
            "employee_no": "R001",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]
    reviewer_detail_resp = client.get(f"/users/{reviewer_id}", headers=_headers(1))
    assert reviewer_detail_resp.status_code == 200
    assert reviewer_detail_resp.json()["id"] == reviewer_id

    employee_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "Alice", "employee_no": "E001", "department": "RD", "roles": ["employee"]},
    )
    assert employee_resp.status_code == 200
    employee_id = employee_resp.json()["id"]

    acceptor_list_resp = client.get("/users/acceptors", headers=_headers(reviewer_id))
    assert acceptor_list_resp.status_code == 200
    assert any(item["id"] == reviewer_id for item in acceptor_list_resp.json())

    acceptor_list_forbidden = client.get("/users/acceptors", headers=_headers(employee_id))
    assert acceptor_list_forbidden.status_code == 403

    me_resp = client.get("/me", headers=_headers(employee_id))
    assert me_resp.status_code == 200
    assert me_resp.json()["id"] == employee_id

    problem_resp = client.post(
        "/problems",
        headers=_headers(employee_id),
        json={
            "title": "构建流水线重复手工操作",
            "scenario": "rd",
            "background": "每次发布都需要重复点选流程",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "单次发布平均浪费20分钟",
            "value_reduce_effort": True,
            "value_statement": "自动化后每周可节省约2小时",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    filtered_problem_resp = client.get(
        "/problems",
        headers=_headers(employee_id),
        params={
            "mine_only": "true",
            "status": "draft",
            "scenario": "rd",
            "created_from": date.today().isoformat(),
            "created_to": date.today().isoformat(),
        },
    )
    assert filtered_problem_resp.status_code == 200
    assert any(item["id"] == problem_id for item in filtered_problem_resp.json())

    task_due_date = (date.today() + timedelta(days=7)).isoformat()
    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "发布自动化脚本",
                "goal": "减少发布手工步骤",
                "scope": "实现发布脚本并接入流水线",
                "due_date": task_due_date,
                "level": "C",
                "reward_total": 600,
                "proposer_ratio": 0.3,
                "accepter_id": reviewer_id,
                "points": 10,
                "badge": "效率之星",
                "acceptance_criteria": [
                    {"description": "脚本稳定运行7天", "type": "quantified"},
                    {"description": "发布步骤由5步降到2步", "type": "behavioral"},
                ],
            },
        },
    )
    assert review_resp.status_code == 200
    task_id = review_resp.json()["id"]

    filtered_task_resp = client.get(
        "/tasks",
        headers=_headers(employee_id),
        params={
            "status": "open",
            "level": "C",
            "scenario": "rd",
            "reward_min": 500,
            "reward_max": 700,
        },
    )
    assert filtered_task_resp.status_code == 200
    task_rows = filtered_task_resp.json()
    target_task = next(item for item in task_rows if item["id"] == task_id)
    assert target_task["scenario"] == "rd"
    assert target_task["active_claim_count"] == 0
    assert target_task["active_claims"] == []

    task_detail_resp = client.get(f"/tasks/{task_id}", headers=_headers(employee_id))
    assert task_detail_resp.status_code == 200
    assert task_detail_resp.json()["id"] == task_id
    assert len(task_detail_resp.json()["acceptance_criteria"]) == 2

    claim_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    assert claim_resp.status_code == 200
    claim_id = claim_resp.json()["claim_id"]

    task_in_progress_resp = client.get(
        "/tasks",
        headers=_headers(employee_id),
        params={"status": "in_progress", "scenario": "rd"},
    )
    assert task_in_progress_resp.status_code == 200
    in_progress_task = next(item for item in task_in_progress_resp.json() if item["id"] == task_id)
    assert in_progress_task["active_claim_count"] >= 1
    claim_snapshot = next(item for item in in_progress_task["active_claims"] if item["claim_id"] == claim_id)
    assert claim_snapshot["mode"] == "individual"
    assert claim_snapshot["status"] == "active"
    assert claim_snapshot["lead_user_id"] == employee_id
    assert claim_snapshot["lead_user_name"] == "Alice"
    assert claim_snapshot["team_size"] == 1
    assert claim_snapshot["created_at"]

    detail_owner_resp = client.get(f"/claims/{claim_id}/detail", headers=_headers(employee_id))
    assert detail_owner_resp.status_code == 200
    assert detail_owner_resp.json()["claim_id"] == claim_id

    detail_acceptor_resp = client.get(f"/claims/{claim_id}/detail", headers=_headers(reviewer_id))
    assert detail_acceptor_resp.status_code == 200

    outsider_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "Outsider", "employee_no": "E999", "department": "RD", "roles": ["employee"]},
    )
    outsider_id = outsider_resp.json()["id"]
    detail_forbidden_resp = client.get(f"/claims/{claim_id}/detail", headers=_headers(outsider_id))
    assert detail_forbidden_resp.status_code == 403

    my_claims_resp = client.get("/claims/mine", headers=_headers(employee_id))
    assert my_claims_resp.status_code == 200
    assert len(my_claims_resp.json()) >= 1

    deliverable_resp = client.post(
        f"/claims/{claim_id}/deliverables",
        headers=_headers(employee_id),
        json={
            "summary": "脚本已上线并接入流水线",
            "evidence_urls": ["https://example.com/screenshot-1"],
            "criteria_results": ["已连续稳定运行7天", "步骤从5步减少为2步"],
        },
    )
    assert deliverable_resp.status_code == 200
    deliverable_id = deliverable_resp.json()["deliverable_id"]

    pending_acceptance_resp = client.get(
        "/deliverables/pending-acceptance/mine",
        headers=_headers(reviewer_id),
    )
    assert pending_acceptance_resp.status_code == 200
    assert any(item["deliverable_id"] == deliverable_id for item in pending_acceptance_resp.json())

    accept_resp = client.post(
        f"/deliverables/{deliverable_id}/accept",
        headers=_headers(reviewer_id),
        json={"result": "approved", "comment": "验收通过"},
    )
    assert accept_resp.status_code == 200
    assert accept_resp.json()["task_status"] == "completed"

    rewards_resp = client.get("/rewards", headers=_headers(employee_id), params={"user_id": employee_id})
    assert rewards_resp.status_code == 200
    rewards = rewards_resp.json()
    assert len(rewards) == 2

    for reward in rewards:
        confirm_resp = client.post(f"/rewards/{reward['id']}/confirm", headers=_headers(reviewer_id))
        assert confirm_resp.status_code == 200
        assert confirm_resp.json()["status"] == "confirmed"

    me_summary_resp = client.get("/me/summary", headers=_headers(employee_id))
    assert me_summary_resp.status_code == 200
    me_summary = me_summary_resp.json()
    assert me_summary["user"]["id"] == employee_id
    assert me_summary["stats"]["total_records"] == 2
    assert me_summary["stats"]["confirmed_records"] == 2
    assert me_summary["stats"]["confirmed_reward_amount"] > 0
    assert me_summary["stats"]["confirmed_points"] >= 10
    assert isinstance(me_summary["badges"], list)
    assert len(me_summary["rewards"]) == 2

    knowledge_resp = client.get("/knowledge", headers=_headers(employee_id))
    assert knowledge_resp.status_code == 200
    knowledge_rows = knowledge_resp.json()
    assert len(knowledge_rows) == 1
    knowledge_id = knowledge_rows[0]["id"]

    knowledge_keyword_resp = client.get(
        "/knowledge",
        headers=_headers(employee_id),
        params={"keyword": "减少发布手工步骤"},
    )
    assert knowledge_keyword_resp.status_code == 200
    assert any(item["id"] == knowledge_id for item in knowledge_keyword_resp.json())

    knowledge_scenario_resp = client.get(
        "/knowledge",
        headers=_headers(employee_id),
        params={"scenario": "rd"},
    )
    assert knowledge_scenario_resp.status_code == 200
    assert any(item["id"] == knowledge_id for item in knowledge_scenario_resp.json())

    knowledge_level_resp = client.get(
        "/knowledge",
        headers=_headers(employee_id),
        params={"level": "C"},
    )
    assert knowledge_level_resp.status_code == 200
    assert any(item["id"] == knowledge_id for item in knowledge_level_resp.json())

    knowledge_recommended_resp = client.get(
        "/knowledge",
        headers=_headers(employee_id),
        params={"recommended": "false"},
    )
    assert knowledge_recommended_resp.status_code == 200
    assert any(item["id"] == knowledge_id for item in knowledge_recommended_resp.json())

    knowledge_detail_resp = client.get(f"/knowledge/{knowledge_id}", headers=_headers(employee_id))
    assert knowledge_detail_resp.status_code == 200
    assert knowledge_detail_resp.json()["id"] == knowledge_id

    dashboard_resp = client.get("/dashboard/overview", headers=_headers(employee_id))
    assert dashboard_resp.status_code == 200
    assert dashboard_resp.json()["task_completed"] == 1
    assert dashboard_resp.json()["task_completion_rate"] > 0
    assert "task_overdue_rate" in dashboard_resp.json()

    app.dependency_overrides.clear()


def test_is_complex_persists_from_review_to_task_reads(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "ComplexReviewer",
            "employee_no": "R960",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    reward_approver_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "ComplexFinance",
            "employee_no": "F960",
            "department": "Finance",
            "roles": ["reward_approver"],
        },
    )
    assert reward_approver_resp.status_code == 200
    reward_approver_id = reward_approver_resp.json()["id"]

    submitter_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "ComplexSubmitter", "employee_no": "E960", "department": "RD", "roles": ["employee"]},
    )
    assert submitter_resp.status_code == 200
    submitter_id = submitter_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(submitter_id),
        json={
            "title": "complex-task-problem",
            "scenario": "rd",
            "background": "complex task wiring",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "persist is_complex from review payload",
            "value_reduce_effort": True,
            "value_statement": "cover task complexity flag",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "complex-task",
                "goal": "persist complexity flag",
                "scope": "single flow",
                "due_date": (date.today() + timedelta(days=4)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "points": 5,
                "is_complex": True,
                "closing_reward_ratio": 0.4,
                "milestones": [
                    {
                        "sequence": 1,
                        "title": "m1",
                        "goal": "first phase",
                        "reward_ratio": 0.3,
                        "acceptance_criteria": [{"description": "phase 1 done", "type": "quantified"}],
                    },
                    {
                        "sequence": 2,
                        "title": "m2",
                        "goal": "second phase",
                        "reward_ratio": 0.3,
                        "acceptance_criteria": [{"description": "phase 2 done", "type": "quantified"}],
                    },
                ],
                "acceptance_criteria": [{"description": "complex flag is saved", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200
    review_payload = review_resp.json()
    assert review_payload["status"] == "approved"

    budget_approve_resp = client.post(
        f"/problems/{problem_id}/budget-review",
        headers=_headers(reward_approver_id),
        json={"approve": True, "comment": "complex task approved"},
    )
    assert budget_approve_resp.status_code == 409
    budget_payload = review_payload
    task_id = budget_payload["id"]
    assert budget_payload["task"]["is_complex"] is True
    assert budget_payload["task"]["task_type"] == "complex"

    task_detail_resp = client.get(f"/tasks/{task_id}", headers=_headers(submitter_id))
    assert task_detail_resp.status_code == 200
    assert task_detail_resp.json()["is_complex"] is True
    assert task_detail_resp.json()["task_type"] == "complex"

    task_list_resp = client.get("/tasks", headers=_headers(submitter_id), params={"status": "open"})
    assert task_list_resp.status_code == 200
    task_row = next(item for item in task_list_resp.json() if item["id"] == task_id)
    assert task_row["is_complex"] is True
    assert task_row["task_type"] == "complex"

    app.dependency_overrides.clear()


def test_release_overdue_claims(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    get_freq_resp = client.get("/system/config/release-overdue-frequency", headers=_headers(1))
    assert get_freq_resp.status_code == 200
    assert get_freq_resp.json()["frequency_minutes"] >= 5

    put_freq_resp = client.put(
        "/system/config/release-overdue-frequency",
        headers=_headers(1),
        json={"frequency_minutes": 15},
    )
    assert put_freq_resp.status_code == 200
    assert put_freq_resp.json()["frequency_minutes"] == 15

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "Reviewer2",
            "employee_no": "R002",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    reviewer_id = reviewer_resp.json()["id"]

    employee_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "Bob", "employee_no": "E002", "department": "RD", "roles": ["employee"]},
    )
    employee_id = employee_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(employee_id),
        json={
            "title": "日志清理流程重复",
            "scenario": "ops",
            "background": "每周人工清理日志",
            "frequency": "weekly",
            "impact_scope": "department",
            "description": "易遗漏导致磁盘告警",
            "value_reduce_effort": True,
            "value_statement": "自动化可减少值班负担",
        },
    )
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "日志清理自动化",
                "goal": "自动化清理",
                "scope": "实现定时清理并留审计日志",
                "due_date": (date.today() - timedelta(days=1)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "acceptance_criteria": [
                    {"description": "定时任务连续运行3天", "type": "quantified"}
                ],
            },
        },
    )
    task_id = review_resp.json()["id"]

    claim_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    assert claim_resp.status_code == 200
    claim_id = claim_resp.json()["claim_id"]

    job_resp = client.post("/jobs/release-overdue", headers=_headers(reviewer_id))
    assert job_resp.status_code == 200
    assert job_resp.json()["released_claims"] >= 1

    claim_activity_resp = client.get(f"/claims/{claim_id}/activities", headers=_headers(employee_id))
    assert claim_activity_resp.status_code == 200
    claim_activity_payload = claim_activity_resp.json()
    assert any(
        item["activity_type"] == "system_event"
        and item["detail"]["event_key"] == "claim_created"
        for item in claim_activity_payload
    )
    assert any(
        item["activity_type"] == "system_event"
        and item["detail"]["event_key"] == "claim_released_overdue"
        for item in claim_activity_payload
    )

    app.dependency_overrides.clear()


def test_claim_overdue_approval_policy(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "ReviewerPolicy",
            "employee_no": "R040",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    employee_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "OverdueUser", "employee_no": "E040", "department": "RD", "roles": ["employee"]},
    )
    assert employee_resp.status_code == 200
    employee_id = employee_resp.json()["id"]

    get_threshold_resp = client.get(
        "/system/config/claim-approval-overdue-threshold",
        headers=_headers(reviewer_id),
    )
    assert get_threshold_resp.status_code == 200
    assert get_threshold_resp.json()["threshold"] >= 1

    set_threshold_resp = client.put(
        "/system/config/claim-approval-overdue-threshold",
        headers=_headers(1),
        json={"threshold": 1},
    )
    assert set_threshold_resp.status_code == 200
    assert set_threshold_resp.json()["threshold"] == 1

    first_problem_resp = client.post(
        "/problems",
        headers=_headers(employee_id),
        json={
            "title": "overdue policy first task",
            "scenario": "ops",
            "background": "prepare overdue count",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "claim first task and let it overdue",
            "value_reduce_effort": True,
            "value_statement": "exercise overdue path",
        },
    )
    assert first_problem_resp.status_code == 200
    first_problem_id = first_problem_resp.json()["id"]

    first_review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        first_problem_id,
        {
            "approve": True,
            "task": {
                "title": "overdue seed task",
                "goal": "seed overdue counter",
                "scope": "single task",
                "due_date": (date.today() - timedelta(days=1)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "acceptance_criteria": [{"description": "create overdue claim", "type": "quantified"}],
            },
        },
    )
    assert first_review_resp.status_code == 200
    first_task_id = first_review_resp.json()["id"]

    first_claim_resp = client.post(
        f"/tasks/{first_task_id}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    assert first_claim_resp.status_code == 200

    release_resp = client.post("/jobs/release-overdue", headers=_headers(reviewer_id))
    assert release_resp.status_code == 200
    assert release_resp.json()["released_claims"] >= 1

    employee_detail_resp = client.get(f"/users/{employee_id}", headers=_headers(1))
    assert employee_detail_resp.status_code == 200
    assert employee_detail_resp.json()["overdue_count"] >= 1

    second_problem_resp = client.post(
        "/problems",
        headers=_headers(employee_id),
        json={
            "title": "overdue policy second task",
            "scenario": "rd",
            "background": "verify approval gate",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "self claim should be blocked after overdue threshold",
            "value_reduce_effort": True,
            "value_statement": "validate approval strategy",
        },
    )
    assert second_problem_resp.status_code == 200
    second_problem_id = second_problem_resp.json()["id"]

    second_review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        second_problem_id,
        {
            "approve": True,
            "task": {
                "title": "approval required task",
                "goal": "enforce approval for overdue users",
                "scope": "single task",
                "due_date": (date.today() + timedelta(days=2)).isoformat(),
                "level": "C",
                "reward_total": 400,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "acceptance_criteria": [{"description": "self claim blocked", "type": "quantified"}],
            },
        },
    )
    assert second_review_resp.status_code == 200
    second_task_id = second_review_resp.json()["id"]

    blocked_claim_resp = client.post(
        f"/tasks/{second_task_id}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    assert blocked_claim_resp.status_code == 403
    assert "approval" in blocked_claim_resp.text.lower()
    my_requests_resp = client.get(
        "/claims/overdue-approvals/mine",
        headers=_headers(employee_id),
        params={"status": "pending"},
    )
    assert my_requests_resp.status_code == 200
    pending_rows = my_requests_resp.json()
    target_request = next(item for item in pending_rows if item["task_id"] == second_task_id)
    request_id = target_request["id"]

    reviewer_pending_resp = client.get(
        "/claims/overdue-approvals/pending",
        headers=_headers(reviewer_id),
        params={"status": "pending"},
    )
    assert reviewer_pending_resp.status_code == 200
    assert any(item["id"] == request_id for item in reviewer_pending_resp.json())

    approve_resp = client.post(
        f"/claims/overdue-approvals/{request_id}/approve",
        headers=_headers(reviewer_id),
        json={"comment": "approved by reviewer"},
    )
    assert approve_resp.status_code == 200
    assert approve_resp.json()["status"] == "approved"
    assert approve_resp.json()["task_id"] == second_task_id

    approved_list_resp = client.get(
        "/claims/overdue-approvals/pending",
        headers=_headers(reviewer_id),
        params={"status": "approved"},
    )
    assert approved_list_resp.status_code == 200
    assert any(item["id"] == request_id for item in approved_list_resp.json())

    third_problem_resp = client.post(
        "/problems",
        headers=_headers(employee_id),
        json={
            "title": "overdue policy third task",
            "scenario": "support",
            "background": "verify reject flow",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "self claim should create request then reviewer rejects",
            "value_reduce_effort": True,
            "value_statement": "validate reject strategy",
        },
    )
    assert third_problem_resp.status_code == 200
    third_problem_id = third_problem_resp.json()["id"]
    third_review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        third_problem_id,
        {
            "approve": True,
            "task": {
                "title": "approval reject task",
                "goal": "reject overdue request",
                "scope": "single task",
                "due_date": (date.today() + timedelta(days=3)).isoformat(),
                "level": "C",
                "reward_total": 500,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "acceptance_criteria": [{"description": "request can be rejected", "type": "quantified"}],
            },
        },
    )
    assert third_review_resp.status_code == 200
    third_task_id = third_review_resp.json()["id"]

    blocked_third_resp = client.post(
        f"/tasks/{third_task_id}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    assert blocked_third_resp.status_code == 403

    third_request_resp = client.get(
        "/claims/overdue-approvals/mine",
        headers=_headers(employee_id),
        params={"status": "pending"},
    )
    assert third_request_resp.status_code == 200
    third_request = next(item for item in third_request_resp.json() if item["task_id"] == third_task_id)
    third_request_id = third_request["id"]

    reject_resp = client.post(
        f"/claims/overdue-approvals/{third_request_id}/reject",
        headers=_headers(reviewer_id),
        json={"comment": "rejected by reviewer"},
    )
    assert reject_resp.status_code == 200
    assert reject_resp.json()["status"] == "rejected"

    app.dependency_overrides.clear()


def test_dashboard_and_export_endpoints(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "ReviewX",
            "employee_no": "R010",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    reviewer_id = reviewer_resp.json()["id"]

    employee_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "DevX",
            "employee_no": "E010",
            "department": "RD",
            "roles": ["employee"],
        },
    )
    employee_id = employee_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(employee_id),
        json={
            "title": "pipeline script optimization",
            "scenario": "rd",
            "background": "manual release is repetitive",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "release requires repeated clicks",
            "value_reduce_effort": True,
            "value_statement": "save team time every week",
        },
    )
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "automate release",
                "goal": "reduce manual release operations",
                "scope": "script + ci integration",
                "due_date": (date.today() + timedelta(days=3)).isoformat(),
                "level": "C",
                "reward_total": 800,
                "proposer_ratio": 0.25,
                "accepter_id": reviewer_id,
                "points": 10,
                "badge": "efficiency-star",
                "acceptance_criteria": [
                    {"description": "script runs stably", "type": "quantified"}
                ],
            },
        },
    )
    task_id = review_resp.json()["id"]

    claim_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    claim_id = claim_resp.json()["claim_id"]

    deliverable_resp = client.post(
        f"/claims/{claim_id}/deliverables",
        headers=_headers(employee_id),
        json={
            "summary": "done",
            "evidence_urls": ["https://example.com/evidence"],
            "criteria_results": ["met"],
        },
    )
    deliverable_id = deliverable_resp.json()["deliverable_id"]

    client.post(
        f"/deliverables/{deliverable_id}/accept",
        headers=_headers(reviewer_id),
        json={"result": "approved", "comment": "ok"},
    )
    rewards_resp = client.get("/rewards", headers=_headers(1))
    reward_id = rewards_resp.json()[0]["id"]
    client.post(f"/rewards/{reward_id}/confirm", headers=_headers(reviewer_id))

    rankings_resp = client.get("/dashboard/rankings", headers=_headers(1), params={"time_range": "all"})
    assert rankings_resp.status_code == 200
    assert "claim_count_ranking" in rankings_resp.json()

    trends_resp = client.get(
        "/dashboard/trends",
        headers=_headers(1),
        params={"time_range": "all", "granularity": "month"},
    )
    assert trends_resp.status_code == 200
    assert len(trends_resp.json()["points"]) >= 1

    dist_resp = client.get("/dashboard/distribution", headers=_headers(1), params={"time_range": "all"})
    assert dist_resp.status_code == 200
    assert "scenario_distribution" in dist_resp.json()

    tasks_export = client.get("/exports/tasks.xlsx", headers=_headers(1))
    assert tasks_export.status_code == 200
    tasks_wb = load_workbook(BytesIO(tasks_export.content))
    assert "Tasks" in tasks_wb.sheetnames

    rewards_export = client.get("/exports/rewards.xlsx", headers=_headers(1))
    assert rewards_export.status_code == 200
    rewards_wb = load_workbook(BytesIO(rewards_export.content))
    assert "Rewards" in rewards_wb.sheetnames

    dashboard_export = client.get(
        "/exports/dashboard.xlsx",
        headers=_headers(1),
        params={"time_range": "all", "granularity": "month"},
    )
    assert dashboard_export.status_code == 200
    dashboard_wb = load_workbook(BytesIO(dashboard_export.content))
    assert "Rankings" in dashboard_wb.sheetnames

    knowledge_pdf = client.get("/exports/knowledge.pdf", headers=_headers(1))
    assert knowledge_pdf.status_code == 200
    assert knowledge_pdf.content.startswith(b"%PDF")

    app.dependency_overrides.clear()


def test_user_status_update_and_disable_effect(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    employee_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "StatusUser",
            "employee_no": "E020",
            "department": "RD",
            "roles": ["employee"],
        },
    )
    assert employee_resp.status_code == 200
    employee_id = employee_resp.json()["id"]

    disable_resp = client.put(
        f"/users/{employee_id}/status",
        headers=_headers(1),
        json={"status": "disabled"},
    )
    assert disable_resp.status_code == 200
    assert disable_resp.json()["status"] == "disabled"

    blocked_problem_resp = client.post(
        "/problems",
        headers=_headers(employee_id),
        json={
            "title": "disabled user should be blocked",
            "scenario": "rd",
            "background": "x",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "x",
            "value_reduce_effort": True,
            "value_statement": "x",
        },
    )
    assert blocked_problem_resp.status_code == 403

    enable_resp = client.put(
        f"/users/{employee_id}/status",
        headers=_headers(1),
        json={"status": "enabled"},
    )
    assert enable_resp.status_code == 200
    assert enable_resp.json()["status"] == "enabled"

    active_users_before_disable = client.get("/users/active", headers=_headers(employee_id))
    assert active_users_before_disable.status_code == 200
    assert any(item["id"] == employee_id for item in active_users_before_disable.json())

    allowed_problem_resp = client.post(
        "/problems",
        headers=_headers(employee_id),
        json={
            "title": "enabled user can submit",
            "scenario": "rd",
            "background": "x",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "x",
            "value_reduce_effort": True,
            "value_statement": "x",
        },
    )
    assert allowed_problem_resp.status_code == 200

    non_admin_update_resp = client.put(
        f"/users/{employee_id}/status",
        headers=_headers(employee_id),
        json={"status": "disabled"},
    )
    assert non_admin_update_resp.status_code == 403

    disable_again_resp = client.put(
        f"/users/{employee_id}/status",
        headers=_headers(1),
        json={"status": "disabled"},
    )
    assert disable_again_resp.status_code == 200

    active_users_after_disable = client.get("/users/active", headers=_headers(1))
    assert active_users_after_disable.status_code == 200
    assert all(item["id"] != employee_id for item in active_users_after_disable.json())

    app.dependency_overrides.clear()


def test_abandon_claim_flow(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "ReviewerAbandon",
            "employee_no": "R030",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    reviewer_id = reviewer_resp.json()["id"]

    employee_a_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "DevA", "employee_no": "E031", "department": "RD", "roles": ["employee"]},
    )
    employee_a_id = employee_a_resp.json()["id"]

    employee_b_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "DevB", "employee_no": "E032", "department": "RD", "roles": ["employee"]},
    )
    employee_b_id = employee_b_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(employee_a_id),
        json={
            "title": "abandon claim case",
            "scenario": "rd",
            "background": "need verify abandon workflow",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "abandon should reopen task when no active claim",
            "value_reduce_effort": True,
            "value_statement": "validate reopen behavior",
        },
    )
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "review abandon behavior",
                "goal": "support abandon operation",
                "scope": "single workflow",
                "due_date": (date.today() + timedelta(days=2)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "acceptance_criteria": [{"description": "flow works", "type": "quantified"}],
            },
        },
    )
    task_id = review_resp.json()["id"]

    claim_a_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(employee_a_id),
        json={"mode": "individual"},
    )
    claim_a_id = claim_a_resp.json()["claim_id"]

    claim_b_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(employee_b_id),
        json={"mode": "individual"},
    )
    claim_b_id = claim_b_resp.json()["claim_id"]

    claim_a_activity_before_resp = client.get(
        f"/claims/{claim_a_id}/activities",
        headers=_headers(employee_a_id),
    )
    assert claim_a_activity_before_resp.status_code == 200
    assert any(
        item["activity_type"] == "system_event"
        and item["detail"]["event_key"] == "claim_created"
        for item in claim_a_activity_before_resp.json()
    )

    forbidden_abandon = client.post(f"/claims/{claim_b_id}/abandon", headers=_headers(employee_a_id))
    assert forbidden_abandon.status_code == 403

    abandon_a_resp = client.post(f"/claims/{claim_a_id}/abandon", headers=_headers(employee_a_id))
    assert abandon_a_resp.status_code == 200
    assert abandon_a_resp.json()["status"] == "abandoned"
    assert abandon_a_resp.json()["task_status"] == "in_progress"

    abandon_b_resp = client.post(f"/claims/{claim_b_id}/abandon", headers=_headers(employee_b_id))
    assert abandon_b_resp.status_code == 200
    assert abandon_b_resp.json()["status"] == "abandoned"
    assert abandon_b_resp.json()["task_status"] == "open"

    claim_b_activity_after_resp = client.get(
        f"/claims/{claim_b_id}/activities",
        headers=_headers(employee_b_id),
    )
    assert claim_b_activity_after_resp.status_code == 200
    assert any(
        item["activity_type"] == "system_event"
        and item["detail"]["event_key"] == "claim_abandoned"
        for item in claim_b_activity_after_resp.json()
    )

    reopen_tasks_resp = client.get("/tasks", headers=_headers(employee_a_id), params={"status": "open"})
    assert reopen_tasks_resp.status_code == 200
    assert any(item["id"] == task_id for item in reopen_tasks_resp.json())

    app.dependency_overrides.clear()


def test_system_config_overview_and_operation_logs(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "LogReviewer",
            "employee_no": "R050",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    overview_resp = client.get("/system/config/overview", headers=_headers(reviewer_id))
    assert overview_resp.status_code == 200
    overview = overview_resp.json()
    assert overview["feishu_sync_frequency_minutes"] >= 5
    assert overview["release_overdue_frequency_minutes"] >= 5
    assert overview["claim_approval_overdue_threshold"] >= 1
    assert "acceptance_templates" in overview

    logs_resp = client.get(
        "/operations/logs",
        headers=_headers(reviewer_id),
        params={"action": "user.create", "limit": 50},
    )
    assert logs_resp.status_code == 200
    rows = logs_resp.json()
    assert any(item["action"] == "user.create" for item in rows)

    app.dependency_overrides.clear()



def test_problem_review_concurrent_guard(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "ReviewGuard",
            "employee_no": "R801",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    employee_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "GuardDev", "employee_no": "E801", "department": "RD", "roles": ["employee"]},
    )
    assert employee_resp.status_code == 200
    employee_id = employee_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(employee_id),
        json={
            "title": "review guard problem",
            "scenario": "rd",
            "background": "guard",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "review once only",
            "value_reduce_effort": True,
            "value_statement": "guard duplicated review",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    first_review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "review guard task",
                "goal": "guard",
                "scope": "single",
                "due_date": (date.today() + timedelta(days=3)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "acceptance_criteria": [{"description": "ok", "type": "quantified"}],
            },
        },
    )
    assert first_review_resp.status_code == 200

    second_review_resp = client.post(
        f"/problems/{problem_id}/review",
        headers=_headers(reviewer_id),
        json={"approve": False, "reject_reason": "second review should fail"},
    )
    assert second_review_resp.status_code == 409

    app.dependency_overrides.clear()


def test_rejected_deliverable_keeps_task_in_progress_with_other_active_claims(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "RejectReviewer",
            "employee_no": "R811",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    owner_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "Owner811", "employee_no": "E811", "department": "RD", "roles": ["employee"]},
    )
    assert owner_resp.status_code == 200
    owner_id = owner_resp.json()["id"]

    peer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "Peer811", "employee_no": "E812", "department": "RD", "roles": ["employee"]},
    )
    assert peer_resp.status_code == 200
    peer_id = peer_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(owner_id),
        json={
            "title": "reject keep progress",
            "scenario": "rd",
            "background": "reject one claim",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "task should remain in progress when another claim is active",
            "value_reduce_effort": True,
            "value_statement": "status transition guard",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "reject status task",
                "goal": "test status",
                "scope": "single",
                "due_date": (date.today() + timedelta(days=3)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "acceptance_criteria": [{"description": "ok", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200
    task_id = review_resp.json()["id"]

    claim_owner_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(owner_id),
        json={"mode": "individual"},
    )
    assert claim_owner_resp.status_code == 200
    claim_owner_id = claim_owner_resp.json()["claim_id"]

    claim_peer_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(peer_id),
        json={"mode": "individual"},
    )
    assert claim_peer_resp.status_code == 200

    deliverable_resp = client.post(
        f"/claims/{claim_owner_id}/deliverables",
        headers=_headers(owner_id),
        json={
            "summary": "to be rejected",
            "criteria_results": ["not enough"],
            "evidence_urls": [],
        },
    )
    assert deliverable_resp.status_code == 200
    deliverable_id = deliverable_resp.json()["deliverable_id"]

    reject_resp = client.post(
        f"/deliverables/{deliverable_id}/accept",
        headers=_headers(reviewer_id),
        json={"result": "rejected", "comment": "reject this claim"},
    )
    assert reject_resp.status_code == 200
    assert reject_resp.json()["task_status"] == "in_progress"

    in_progress_resp = client.get("/tasks", headers=_headers(owner_id), params={"status": "in_progress"})
    assert in_progress_resp.status_code == 200
    assert any(item["id"] == task_id for item in in_progress_resp.json())

    app.dependency_overrides.clear()


def test_release_overdue_boundary_due_today_is_not_overdue(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "BoundaryReviewer",
            "employee_no": "R821",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    employee_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "BoundaryDev", "employee_no": "E821", "department": "RD", "roles": ["employee"]},
    )
    assert employee_resp.status_code == 200
    employee_id = employee_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(employee_id),
        json={
            "title": "overdue boundary problem",
            "scenario": "ops",
            "background": "boundary",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "due today should not be overdue",
            "value_reduce_effort": True,
            "value_statement": "boundary test",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "boundary task",
                "goal": "boundary",
                "scope": "single",
                "due_date": date.today().isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "acceptance_criteria": [{"description": "ok", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200
    task_id = review_resp.json()["id"]

    claim_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    assert claim_resp.status_code == 200

    release_resp = client.post("/jobs/release-overdue", headers=_headers(reviewer_id))
    assert release_resp.status_code == 200
    assert release_resp.json()["released_claims"] == 0

    claims_resp = client.get("/claims/mine", headers=_headers(employee_id), params={"status": "active"})
    assert claims_resp.status_code == 200
    assert any(item["task_id"] == task_id for item in claims_resp.json())

    app.dependency_overrides.clear()


def test_pagination_and_like_escape_for_lists(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "PagerReviewer",
            "employee_no": "R831",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    employee_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "PagerDev", "employee_no": "E831", "department": "RD", "roles": ["employee"]},
    )
    assert employee_resp.status_code == 200
    employee_id = employee_resp.json()["id"]

    created_problem_ids: list[int] = []
    created_task_ids: list[int] = []
    for idx in range(1, 3):
        problem_resp = client.post(
            "/problems",
            headers=_headers(employee_id),
            json={
                "title": f"pager problem {idx}",
                "scenario": "rd",
                "background": "pager",
                "frequency": "weekly",
                "impact_scope": "team",
                "description": "pager problem",
                "value_reduce_effort": True,
                "value_statement": "pager value",
            },
        )
        assert problem_resp.status_code == 200
        problem_id = problem_resp.json()["id"]
        created_problem_ids.append(problem_id)

        review_resp = _review_problem_and_create_task(
            client,
            reviewer_id,
            problem_id,
            {
                "approve": True,
                "task": {
                    "title": f"pager task {idx}",
                    "goal": "pager",
                    "scope": "single",
                    "due_date": (date.today() + timedelta(days=3)).isoformat(),
                    "level": "C",
                    "reward_total": 300,
                    "proposer_ratio": 0.2,
                    "accepter_id": reviewer_id,
                    "acceptance_criteria": [{"description": "ok", "type": "quantified"}],
                },
            },
        )
        assert review_resp.status_code == 200
        created_task_ids.append(review_resp.json()["id"])

    problems_page_1 = client.get(
        "/problems",
        headers=_headers(employee_id),
        params={"mine_only": "true", "limit": 1, "offset": 0},
    )
    assert problems_page_1.status_code == 200
    assert len(problems_page_1.json()) == 1

    problems_page_2 = client.get(
        "/problems",
        headers=_headers(employee_id),
        params={"mine_only": "true", "limit": 1, "offset": 1},
    )
    assert problems_page_2.status_code == 200
    assert len(problems_page_2.json()) == 1
    assert problems_page_1.json()[0]["id"] != problems_page_2.json()[0]["id"]

    tasks_page_1 = client.get("/tasks", headers=_headers(employee_id), params={"limit": 1, "offset": 0})
    assert tasks_page_1.status_code == 200
    assert len(tasks_page_1.json()) == 1
    tasks_page_2 = client.get("/tasks", headers=_headers(employee_id), params={"limit": 1, "offset": 1})
    assert tasks_page_2.status_code == 200
    assert len(tasks_page_2.json()) == 1
    assert tasks_page_1.json()[0]["id"] != tasks_page_2.json()[0]["id"]

    claim_resp = client.post(
        f"/tasks/{created_task_ids[0]}/claims",
        headers=_headers(employee_id),
        json={
            "mode": "team",
            "lead_user_id": employee_id,
            "members": [
                {"user_id": employee_id, "ratio": 0.3333},
                {"user_id": reviewer_id, "ratio": 0.3333},
                {"user_id": 1, "ratio": 0.3334},
            ],
        },
    )
    assert claim_resp.status_code == 200
    claim_id = claim_resp.json()["claim_id"]

    deliverable_resp = client.post(
        f"/claims/{claim_id}/deliverables",
        headers=_headers(employee_id),
        json={
            "summary": "pager deliverable",
            "criteria_results": ["ok"],
            "evidence_urls": [],
        },
    )
    assert deliverable_resp.status_code == 200
    deliverable_id = deliverable_resp.json()["deliverable_id"]

    accept_resp = client.post(
        f"/deliverables/{deliverable_id}/accept",
        headers=_headers(reviewer_id),
        json={"result": "approved", "comment": "ok"},
    )
    assert accept_resp.status_code == 200

    rewards_page = client.get("/rewards", headers=_headers(employee_id), params={"limit": 1, "offset": 0})
    assert rewards_page.status_code == 200
    assert len(rewards_page.json()) == 1

    all_rewards = client.get("/rewards", headers=_headers(employee_id))
    assert all_rewards.status_code == 200
    related_rewards = [item for item in all_rewards.json() if item["task_id"] == created_task_ids[0]]
    assert len(related_rewards) == 4
    assert round(sum(item["amount"] for item in related_rewards), 2) == 300.00

    escaped_keyword_resp = client.get("/knowledge", headers=_headers(employee_id), params={"keyword": "%"})
    assert escaped_keyword_resp.status_code == 200
    assert escaped_keyword_resp.json() == []

    app.dependency_overrides.clear()


def test_rejected_problem_can_be_modified_and_resubmitted(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "ResubmitReviewer",
            "employee_no": "R901",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    submitter_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "ResubmitDev", "employee_no": "E901", "department": "RD", "roles": ["employee"]},
    )
    assert submitter_resp.status_code == 200
    submitter_id = submitter_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(submitter_id),
        json={
            "title": "resubmit problem",
            "scenario": "rd",
            "background": "original background",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "original description",
            "value_reduce_effort": True,
            "value_statement": "original value",
            "task_draft": {
                "goal": "original goal",
                "scope": "original scope",
                "due_date": (date.today() + timedelta(days=5)).isoformat(),
                "acceptance_criteria": [{"description": "original ok", "type": "quantified"}],
                "self_reflection": "original reflection",
            },
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    submit_resp = client.post(
        f"/problems/{problem_id}/submit-for-review",
        headers=_headers(submitter_id),
    )
    assert submit_resp.status_code == 200

    reject_resp = client.post(
        f"/problems/{problem_id}/review",
        headers=_headers(reviewer_id),
        json={"approve": False, "reject_reason": "need clearer statement"},
    )
    assert reject_resp.status_code == 200
    assert reject_resp.json() is None

    outsider_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "ResubmitOutsider", "employee_no": "E903", "department": "RD", "roles": ["employee"]},
    )
    assert outsider_resp.status_code == 200
    outsider_id = outsider_resp.json()["id"]

    forbidden_resubmit_resp = client.put(
        f"/problems/{problem_id}/resubmit",
        headers=_headers(outsider_id),
        json={
            "title": "blocked resubmit",
            "scenario": "rd",
            "background": "blocked",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "blocked",
            "value_reduce_effort": True,
            "value_statement": "blocked",
        },
    )
    assert forbidden_resubmit_resp.status_code == 403

    resubmit_resp = client.put(
        f"/problems/{problem_id}/resubmit",
        headers=_headers(submitter_id),
        json={
            "title": "resubmitted problem",
            "scenario": "ops",
            "background": "updated background",
            "frequency": "monthly",
            "impact_scope": "department",
            "description": "updated description",
            "value_reduce_effort": True,
            "value_reduce_cost": True,
            "value_statement": "updated value",
            "task_draft": {
                "goal": "updated goal",
                "scope": "updated scope",
                "due_date": (date.today() + timedelta(days=7)).isoformat(),
                "acceptance_criteria": [{"description": "updated ok", "type": "quantified"}],
                "self_reflection": "updated reflection",
            },
        },
    )
    assert resubmit_resp.status_code == 200
    assert resubmit_resp.json()["status"] == "draft"
    assert resubmit_resp.json()["reject_reason"] is None

    detail_resp = client.get(f"/problems/{problem_id}", headers=_headers(submitter_id))
    assert detail_resp.status_code == 200
    assert detail_resp.json()["title"] == "resubmitted problem"
    assert detail_resp.json()["scenario"] == "ops"
    assert detail_resp.json()["frequency"] == "monthly"
    assert detail_resp.json()["status"] == "draft"

    submit_resubmitted_resp = client.post(
        f"/problems/{problem_id}/submit-for-review",
        headers=_headers(submitter_id),
    )
    assert submit_resubmitted_resp.status_code == 200
    assert submit_resubmitted_resp.json()["status"] == "pending_review"

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "resubmitted task",
                "goal": "approve after resubmit",
                "scope": "single",
                "due_date": (date.today() + timedelta(days=5)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "acceptance_criteria": [{"description": "ok", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200

    app.dependency_overrides.clear()


def test_claim_limit_two_active_claims_per_user(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "ClaimLimitReviewer",
            "employee_no": "R902",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    employee_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "ClaimLimitDev", "employee_no": "E902", "department": "RD", "roles": ["employee"]},
    )
    assert employee_resp.status_code == 200
    employee_id = employee_resp.json()["id"]

    task_ids: list[int] = []
    for idx in range(1, 4):
        problem_resp = client.post(
            "/problems",
            headers=_headers(employee_id),
            json={
                "title": f"claim limit problem {idx}",
                "scenario": "rd",
                "background": "claim limit",
                "frequency": "weekly",
                "impact_scope": "team",
                "description": "prepare claim limit tasks",
                "value_reduce_effort": True,
                "value_statement": "claim limit validation",
            },
        )
        assert problem_resp.status_code == 200
        problem_id = problem_resp.json()["id"]

        review_resp = _review_problem_and_create_task(
            client,
            reviewer_id,
            problem_id,
            {
                "approve": True,
                "task": {
                    "title": f"claim limit task {idx}",
                    "goal": "verify max active claims",
                    "scope": "single",
                    "due_date": (date.today() + timedelta(days=7)).isoformat(),
                    "level": "C",
                    "reward_total": 300,
                    "proposer_ratio": 0.2,
                    "accepter_id": reviewer_id,
                    "acceptance_criteria": [{"description": "ok", "type": "quantified"}],
                },
            },
        )
        assert review_resp.status_code == 200
        task_ids.append(review_resp.json()["id"])

    first_claim_resp = client.post(
        f"/tasks/{task_ids[0]}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    assert first_claim_resp.status_code == 200
    first_claim_id = first_claim_resp.json()["claim_id"]

    second_claim_resp = client.post(
        f"/tasks/{task_ids[1]}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    assert second_claim_resp.status_code == 200

    third_claim_resp = client.post(
        f"/tasks/{task_ids[2]}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    assert third_claim_resp.status_code == 400
    assert "最多进行2个揭榜" in third_claim_resp.text

    abandon_resp = client.post(f"/claims/{first_claim_id}/abandon", headers=_headers(employee_id))
    assert abandon_resp.status_code == 200

    retry_third_claim_resp = client.post(
        f"/tasks/{task_ids[2]}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    assert retry_third_claim_resp.status_code == 200

    app.dependency_overrides.clear()

def test_accept_deliverable_rejects_when_already_finalized(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "AcceptStateReviewer",
            "employee_no": "R950",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    employee_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "AcceptStateDev", "employee_no": "E950", "department": "RD", "roles": ["employee"]},
    )
    assert employee_resp.status_code == 200
    employee_id = employee_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(employee_id),
        json={
            "title": "accept-state-problem",
            "scenario": "rd",
            "background": "state guard",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "verify accept status guard",
            "value_reduce_effort": True,
            "value_statement": "prevent repeated acceptance",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "accept-state-task",
                "goal": "validate accept status guard",
                "scope": "single",
                "due_date": (date.today() + timedelta(days=3)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "acceptance_criteria": [{"description": "ok", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200
    task_id = review_resp.json()["id"]

    claim_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    assert claim_resp.status_code == 200
    claim_id = claim_resp.json()["claim_id"]

    deliverable_resp = client.post(
        f"/claims/{claim_id}/deliverables",
        headers=_headers(employee_id),
        json={"summary": "done", "criteria_results": ["ok"], "evidence_urls": []},
    )
    assert deliverable_resp.status_code == 200
    deliverable_id = deliverable_resp.json()["deliverable_id"]

    first_accept_resp = client.post(
        f"/deliverables/{deliverable_id}/accept",
        headers=_headers(reviewer_id),
        json={"result": "approved", "comment": "ok"},
    )
    assert first_accept_resp.status_code == 200

    second_accept_resp = client.post(
        f"/deliverables/{deliverable_id}/accept",
        headers=_headers(reviewer_id),
        json={"result": "rejected", "comment": "should be blocked"},
    )
    assert second_accept_resp.status_code == 400

    app.dependency_overrides.clear()


def test_acceptor_cannot_accept_own_deliverable(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    worker_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "SelfAcceptUser",
            "employee_no": "E951",
            "department": "RD",
            "roles": ["employee", "acceptor"],
        },
    )
    assert worker_resp.status_code == 200
    worker_id = worker_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(worker_id),
        json={
            "title": "self-accept-problem",
            "scenario": "rd",
            "background": "self accept check",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "self accept should be forbidden",
            "value_reduce_effort": True,
            "value_statement": "conflict check",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        1,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "self-accept-task",
                "goal": "self accept forbidden",
                "scope": "single",
                "due_date": (date.today() + timedelta(days=3)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": worker_id,
                "acceptance_criteria": [{"description": "ok", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200
    task_id = review_resp.json()["id"]

    claim_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(worker_id),
        json={"mode": "individual"},
    )
    assert claim_resp.status_code == 200
    claim_id = claim_resp.json()["claim_id"]

    deliverable_resp = client.post(
        f"/claims/{claim_id}/deliverables",
        headers=_headers(worker_id),
        json={"summary": "done", "criteria_results": ["ok"], "evidence_urls": []},
    )
    assert deliverable_resp.status_code == 200
    deliverable_id = deliverable_resp.json()["deliverable_id"]

    accept_resp = client.post(
        f"/deliverables/{deliverable_id}/accept",
        headers=_headers(worker_id),
        json={"result": "approved", "comment": "self"},
    )
    assert accept_resp.status_code == 403

    app.dependency_overrides.clear()


def test_overdue_claim_can_be_abandoned_and_count_reduced(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "OverdueFixReviewer",
            "employee_no": "R952",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    employee_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "OverdueFixDev", "employee_no": "E952", "department": "RD", "roles": ["employee"]},
    )
    assert employee_resp.status_code == 200
    employee_id = employee_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(employee_id),
        json={
            "title": "overdue-fix-problem",
            "scenario": "ops",
            "background": "overdue recovery",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "allow overdue claim to exit",
            "value_reduce_effort": True,
            "value_statement": "overdue should not deadlock",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "overdue-fix-task",
                "goal": "make claim overdue",
                "scope": "single",
                "due_date": (date.today() - timedelta(days=1)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "acceptance_criteria": [{"description": "ok", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200
    task_id = review_resp.json()["id"]

    claim_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(employee_id),
        json={"mode": "individual"},
    )
    assert claim_resp.status_code == 200
    claim_id = claim_resp.json()["claim_id"]

    release_resp = client.post("/jobs/release-overdue", headers=_headers(reviewer_id))
    assert release_resp.status_code == 200
    assert release_resp.json()["released_claims"] >= 1

    before_resp = client.get(f"/users/{employee_id}", headers=_headers(1))
    assert before_resp.status_code == 200
    assert before_resp.json()["overdue_count"] >= 1

    abandon_resp = client.post(f"/claims/{claim_id}/abandon", headers=_headers(employee_id))
    assert abandon_resp.status_code == 200
    assert abandon_resp.json()["status"] == "abandoned"

    after_resp = client.get(f"/users/{employee_id}", headers=_headers(1))
    assert after_resp.status_code == 200
    assert after_resp.json()["overdue_count"] == max(0, before_resp.json()["overdue_count"] - 1)

    app.dependency_overrides.clear()


def test_submitter_cannot_review_own_problem(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    problem_resp = client.post(
        "/problems",
        headers=_headers(1),
        json={
            "title": "self-review-problem",
            "scenario": "rd",
            "background": "self review check",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "submitter should not review own problem",
            "value_reduce_effort": True,
            "value_statement": "conflict check",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = client.post(
        f"/problems/{problem_id}/review",
        headers=_headers(1),
        json={
            "approve": True,
            "task": {
                "title": "self-review-task",
                "goal": "forbidden",
                "scope": "single",
                "due_date": (date.today() + timedelta(days=3)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": 1,
                "acceptance_criteria": [{"description": "ok", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 403

    app.dependency_overrides.clear()


def test_incomplete_problem_draft_requires_submit_for_review(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "DraftGateReviewer",
            "employee_no": "R990",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    submitter_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "DraftGateSubmitter", "employee_no": "E990", "department": "RD", "roles": ["employee"]},
    )
    assert submitter_resp.status_code == 200
    submitter_id = submitter_resp.json()["id"]

    create_resp = client.post(
        "/problems",
        headers=_headers(submitter_id),
        json={
            "title": "incomplete-draft-problem",
            "scenario": "rd",
            "background": "draft should stay draft",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "missing task draft should not enter reviewer queue",
            "value_reduce_effort": True,
            "value_statement": "protect reviewer queue quality",
        },
    )
    assert create_resp.status_code == 200
    problem_id = create_resp.json()["id"]
    assert create_resp.json()["status"] == "draft"

    submit_resp = client.post(
        f"/problems/{problem_id}/submit-for-review",
        headers=_headers(submitter_id),
    )
    assert submit_resp.status_code == 400
    assert "task draft is incomplete" in submit_resp.text

    resubmit_resp = client.put(
        f"/problems/{problem_id}/resubmit",
        headers=_headers(submitter_id),
        json={
            "title": "incomplete-draft-problem-updated",
            "scenario": "rd",
            "background": "still incomplete",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "resubmit should also stay draft",
            "value_reduce_effort": True,
            "value_statement": "same issue",
        },
    )
    assert resubmit_resp.status_code == 200
    assert resubmit_resp.json()["status"] == "draft"

    detail_resp = client.get(f"/problems/{problem_id}", headers=_headers(submitter_id))
    assert detail_resp.status_code == 200
    assert detail_resp.json()["status"] == "draft"

    complete_resubmit_resp = client.put(
        f"/problems/{problem_id}/resubmit",
        headers=_headers(submitter_id),
        json={
            "title": "complete-draft-problem",
            "scenario": "rd",
            "background": "ready for review",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "complete task draft can be submitted later",
            "value_reduce_effort": True,
            "value_statement": "ready",
            "task_draft": {
                "goal": "ship a valid task draft",
                "scope": "single workflow",
                "due_date": (date.today() + timedelta(days=5)).isoformat(),
                "acceptance_criteria": [{"description": "works", "type": "quantified"}],
                "self_reflection": "completed required fields",
            },
        },
    )
    assert complete_resubmit_resp.status_code == 200
    assert complete_resubmit_resp.json()["status"] == "draft"

    complete_submit_resp = client.post(
        f"/problems/{problem_id}/submit-for-review",
        headers=_headers(submitter_id),
    )
    assert complete_submit_resp.status_code == 200
    assert complete_submit_resp.json()["status"] == "pending_review"

    reviewer_queue_resp = client.get(
        "/problems",
        headers=_headers(reviewer_id),
        params={"status": "pending_review"},
    )
    assert reviewer_queue_resp.status_code == 200
    assert any(item["id"] == problem_id for item in reviewer_queue_resp.json())

    app.dependency_overrides.clear()


def test_deliverable_rework_has_max_attempts(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "ReworkReviewer",
            "employee_no": "R980",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    submitter_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "ReworkSubmitter", "employee_no": "E980", "department": "OPS", "roles": ["employee"]},
    )
    assert submitter_resp.status_code == 200
    submitter_id = submitter_resp.json()["id"]

    executor_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "ReworkExecutor", "employee_no": "E981", "department": "RD", "roles": ["employee"]},
    )
    assert executor_resp.status_code == 200
    executor_id = executor_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(submitter_id),
        json={
            "title": "rework-limit-problem",
            "scenario": "ops",
            "background": "rework limit",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "limit rework attempts",
            "value_reduce_effort": True,
            "value_statement": "prevent endless loop",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "rework-limit-task",
                "goal": "validate rework cap",
                "scope": "single",
                "due_date": (date.today() + timedelta(days=2)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "acceptance_criteria": [{"description": "ok", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200
    task_id = review_resp.json()["id"]

    claim_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(executor_id),
        json={"mode": "individual"},
    )
    assert claim_resp.status_code == 200
    claim_id = claim_resp.json()["claim_id"]

    deliverable_resp = client.post(
        f"/claims/{claim_id}/deliverables",
        headers=_headers(executor_id),
        json={"summary": "round1", "criteria_results": ["ok"], "evidence_urls": []},
    )
    assert deliverable_resp.status_code == 200
    deliverable_id = deliverable_resp.json()["deliverable_id"]

    for round_no in range(3):
        accept_resp = client.post(
            f"/deliverables/{deliverable_id}/accept",
            headers=_headers(reviewer_id),
            json={"result": "rework", "comment": f"round-{round_no + 1}"},
        )
        assert accept_resp.status_code == 200
        resubmit_resp = client.post(
            f"/claims/{claim_id}/deliverables",
            headers=_headers(executor_id),
            json={"summary": f"round{round_no + 2}", "criteria_results": ["ok"], "evidence_urls": []},
        )
        assert resubmit_resp.status_code == 200

    blocked_resp = client.post(
        f"/deliverables/{deliverable_id}/accept",
        headers=_headers(reviewer_id),
        json={"result": "rework", "comment": "round-4"},
    )
    assert blocked_resp.status_code == 400
    assert "max rework attempts reached" in blocked_resp.text

    app.dependency_overrides.clear()


def test_pricing_review_requires_completed_analysis(tmp_path: Path, monkeypatch) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "AnalysisGateReviewer",
            "employee_no": "R981",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    submitter_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "AnalysisGateSubmitter", "employee_no": "E982", "department": "RD", "roles": ["employee"]},
    )
    assert submitter_resp.status_code == 200
    submitter_id = submitter_resp.json()["id"]

    create_resp = client.post(
        "/problems",
        headers=_headers(submitter_id),
        json={
            "title": "analysis-required-problem",
            "scenario": "rd",
            "background": "analysis gate",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "approval should wait for analysis",
            "value_reduce_effort": True,
            "value_statement": "quality gate",
            "task_draft": {
                "goal": "ship mvp",
                "scope": "single flow",
                "due_date": (date.today() + timedelta(days=5)).isoformat(),
                "acceptance_criteria": [{"description": "works", "type": "quantified"}],
                "self_reflection": "validated by submitter",
            },
        },
    )
    assert create_resp.status_code == 200
    problem_id = create_resp.json()["id"]

    submit_resp = client.post(
        f"/problems/{problem_id}/submit-for-review",
        headers=_headers(submitter_id),
    )
    assert submit_resp.status_code == 200

    def fake_trigger(problem_id: int) -> None:
        return None

    monkeypatch.setattr("app.routers.problems._trigger_analysis_background", fake_trigger)
    analyze_resp = client.post(
        f"/problems/{problem_id}/analyze",
        headers=_headers(submitter_id),
    )
    assert analyze_resp.status_code == 200

    review_resp = client.post(
        f"/problems/{problem_id}/review",
        headers=_headers(reviewer_id),
        json={
            "approve": True,
            "pricing": {
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "points": 5,
                "badge": None,
            },
        },
    )
    assert review_resp.status_code == 409
    assert "analysis is still running" in review_resp.text

    app.dependency_overrides.clear()


def test_task_activity_timeline_permissions(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "ActivityReviewer",
            "employee_no": "R970",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    claimant_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "ActivityLead", "employee_no": "E970", "department": "RD", "roles": ["employee"]},
    )
    assert claimant_resp.status_code == 200
    claimant_id = claimant_resp.json()["id"]

    viewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "ActivityViewer", "employee_no": "E971", "department": "OPS", "roles": ["employee"]},
    )
    assert viewer_resp.status_code == 200
    viewer_id = viewer_resp.json()["id"]

    outsider_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "ActivityOutsider", "employee_no": "E972", "department": "OPS", "roles": ["employee"]},
    )
    assert outsider_resp.status_code == 200
    outsider_id = outsider_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(claimant_id),
        json={
            "title": "task-activity-problem",
            "scenario": "rd",
            "background": "add timeline coverage",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "verify task activity permissions",
            "value_reduce_effort": True,
            "value_statement": "cover timeline behavior",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "task-activity-task",
                "goal": "support comments and progress updates",
                "scope": "timeline permissions only",
                "due_date": (date.today() + timedelta(days=5)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "points": 5,
                "acceptance_criteria": [{"description": "timeline works", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200
    task_id = review_resp.json()["id"]

    task_detail_resp = client.get(f"/tasks/{task_id}", headers=_headers(viewer_id))
    assert task_detail_resp.status_code == 200
    assert task_detail_resp.json()["is_complex"] is False

    empty_list_resp = client.get(f"/tasks/{task_id}/activities", headers=_headers(viewer_id))
    assert empty_list_resp.status_code == 200
    assert empty_list_resp.json() == []

    blank_comment_resp = client.post(
        f"/tasks/{task_id}/activities",
        headers=_headers(viewer_id),
        json={"activity_type": "comment", "content": "   "},
    )
    assert blank_comment_resp.status_code == 422

    comment_resp = client.post(
        f"/tasks/{task_id}/activities",
        headers=_headers(viewer_id),
        json={"activity_type": "comment", "content": "I can help review this."},
    )
    assert comment_resp.status_code == 200
    assert comment_resp.json()["task_id"] == task_id
    assert comment_resp.json()["claim_id"] is None
    assert comment_resp.json()["activity_type"] == "comment"
    assert comment_resp.json()["actor_user_id"] == viewer_id
    assert comment_resp.json()["actor_user_name"] == "ActivityViewer"
    assert comment_resp.json()["claim_name"] is None
    assert comment_resp.json()["content"] == "I can help review this."
    assert comment_resp.json()["attachment_urls"] == []
    assert comment_resp.json()["detail"] == {}

    outsider_task_activity_resp = client.get(f"/tasks/{task_id}/activities", headers=_headers(outsider_id))
    assert outsider_task_activity_resp.status_code == 200
    outsider_task_activity_payload = outsider_task_activity_resp.json()
    assert len(outsider_task_activity_payload) == 1
    assert outsider_task_activity_payload[0]["id"] == comment_resp.json()["id"]

    claim_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(claimant_id),
        json={"mode": "individual"},
    )
    assert claim_resp.status_code == 200
    claim_id = claim_resp.json()["claim_id"]

    claim_created_activity_resp = client.get(f"/claims/{claim_id}/activities", headers=_headers(claimant_id))
    assert claim_created_activity_resp.status_code == 200
    claim_created_activity_payload = claim_created_activity_resp.json()
    assert len(claim_created_activity_payload) == 1
    claim_created_event = claim_created_activity_payload[0]
    assert claim_created_event["activity_type"] == "system_event"
    assert claim_created_event["actor_user_name"] == "ActivityLead"
    assert claim_created_event["claim_name"] == "个人 · ActivityLead"
    assert claim_created_event["detail"]["event_key"] == "claim_created"

    active_claim_options_resp = client.get(f"/tasks/{task_id}/claims/active", headers=_headers(claimant_id))
    assert active_claim_options_resp.status_code == 200
    active_claim_options = active_claim_options_resp.json()
    assert len(active_claim_options) == 1
    assert active_claim_options[0]["claim_id"] == claim_id
    assert active_claim_options[0]["lead_user_name"] == "ActivityLead"

    outsider_active_claim_options_resp = client.get(f"/tasks/{task_id}/claims/active", headers=_headers(outsider_id))
    assert outsider_active_claim_options_resp.status_code == 200
    assert outsider_active_claim_options_resp.json() == []

    note_resp = client.post(
        f"/tasks/{task_id}/activities",
        headers=_headers(reviewer_id),
        json={
            "activity_type": "official_note",
            "content": "Please keep updates in this timeline.",
            "detail": {"audience": "all"},
        },
    )
    assert note_resp.status_code == 200
    assert note_resp.json()["activity_type"] == "official_note"
    assert note_resp.json()["claim_id"] is None
    assert note_resp.json()["detail"] == {"audience": "all"}

    viewer_note_resp = client.post(
        f"/tasks/{task_id}/activities",
        headers=_headers(viewer_id),
        json={"activity_type": "official_note", "content": "Not allowed"},
    )
    assert viewer_note_resp.status_code == 403

    blocker_attachment_resp = client.post(
        "/attachments/upload",
        headers=_headers(claimant_id),
        files={"file": ("blocker.txt", b"need access", "text/plain")},
    )
    assert blocker_attachment_resp.status_code == 200
    blocker_attachment_id = blocker_attachment_resp.json()["id"]

    blocker_resp = client.post(
        f"/tasks/{task_id}/activities",
        headers=_headers(claimant_id),
        json={
            "activity_type": "blocker",
            "claim_id": claim_id,
            "content": "Blocked on test environment access.",
            "detail": {"severity": "high"},
            "attachment_ids": [blocker_attachment_id],
        },
    )
    assert blocker_resp.status_code == 200
    assert blocker_resp.json()["task_id"] == task_id
    assert blocker_resp.json()["claim_id"] == claim_id
    assert blocker_resp.json()["claim_name"] == "个人 · ActivityLead"
    assert blocker_resp.json()["activity_type"] == "blocker"
    assert blocker_resp.json()["detail"] == {"severity": "high"}
    assert blocker_resp.json()["attachment_urls"] == [f"/attachments/{blocker_attachment_id}/download"]

    blocker_download_resp = client.get(
        blocker_resp.json()["attachment_urls"][0],
        headers=_headers(claimant_id),
    )
    assert blocker_download_resp.status_code == 200
    assert blocker_download_resp.content == b"need access"

    progress_resp = client.post(
        f"/tasks/{task_id}/activities",
        headers=_headers(claimant_id),
        json={
            "activity_type": "progress_update",
            "claim_id": claim_id,
            "content": "Finished the first draft.",
        },
    )
    assert progress_resp.status_code == 200
    assert progress_resp.json()["task_id"] == task_id
    assert progress_resp.json()["claim_id"] == claim_id
    assert progress_resp.json()["activity_type"] == "progress_update"
    assert progress_resp.json()["actor_user_id"] == claimant_id

    reviewer_progress_resp = client.post(
        f"/tasks/{task_id}/activities",
        headers=_headers(reviewer_id),
        json={
            "activity_type": "progress_update",
            "claim_id": claim_id,
            "content": "Reviewer confirmed the first draft is in progress.",
        },
    )
    assert reviewer_progress_resp.status_code == 200
    assert reviewer_progress_resp.json()["claim_id"] == claim_id
    assert reviewer_progress_resp.json()["actor_user_id"] == reviewer_id

    outsider_task_activity_after_progress_resp = client.get(
        f"/tasks/{task_id}/activities",
        headers=_headers(outsider_id),
    )
    assert outsider_task_activity_after_progress_resp.status_code == 200
    outsider_task_activity_after_progress_payload = outsider_task_activity_after_progress_resp.json()
    assert {item["id"] for item in outsider_task_activity_after_progress_payload} == {
        comment_resp.json()["id"],
        note_resp.json()["id"],
    }

    outsider_blocker_resp = client.get(
        blocker_resp.json()["attachment_urls"][0],
        headers=_headers(outsider_id),
    )
    assert outsider_blocker_resp.status_code == 403

    system_event_create_resp = client.post(
        f"/tasks/{task_id}/activities",
        headers=_headers(reviewer_id),
        json={"activity_type": "system_event", "content": "should fail"},
    )
    assert system_event_create_resp.status_code == 403

    viewer_progress_resp = client.post(
        f"/tasks/{task_id}/activities",
        headers=_headers(viewer_id),
        json={"activity_type": "progress_update", "content": "Pretending to make progress."},
    )
    assert viewer_progress_resp.status_code == 403

    claim_activity_resp = client.get(f"/claims/{claim_id}/activities", headers=_headers(claimant_id))
    assert claim_activity_resp.status_code == 200
    claim_activity_payload = claim_activity_resp.json()
    assert {item["id"] for item in claim_activity_payload} == {
        claim_created_event["id"],
        blocker_resp.json()["id"],
        progress_resp.json()["id"],
        reviewer_progress_resp.json()["id"],
    }
    assert any(
        item["activity_type"] == "system_event"
        and item["detail"]["event_key"] == "claim_created"
        for item in claim_activity_payload
    )

    deliverable_resp = client.post(
        f"/claims/{claim_id}/deliverables",
        headers=_headers(claimant_id),
        json={
            "summary": "Delivered the first working version.",
            "criteria_results": ["timeline works"],
            "evidence_urls": [],
        },
    )
    assert deliverable_resp.status_code == 200
    deliverable_id = deliverable_resp.json()["deliverable_id"]

    accept_resp = client.post(
        f"/deliverables/{deliverable_id}/accept",
        headers=_headers(reviewer_id),
        json={"result": "approved", "comment": "Looks good."},
    )
    assert accept_resp.status_code == 200
    assert accept_resp.json()["task_status"] == "completed"

    claim_activity_after_accept_resp = client.get(
        f"/claims/{claim_id}/activities",
        headers=_headers(claimant_id),
    )
    assert claim_activity_after_accept_resp.status_code == 200
    claim_activity_after_accept_payload = claim_activity_after_accept_resp.json()
    assert any(
        item["activity_type"] == "system_event"
        and item["detail"]["event_key"] == "deliverable_submitted"
        for item in claim_activity_after_accept_payload
    )
    assert any(
        item["activity_type"] == "system_event"
        and item["detail"]["event_key"] == "deliverable_approved"
        for item in claim_activity_after_accept_payload
    )

    delete_comment_resp = client.request(
        "DELETE",
        f"/activities/{comment_resp.json()['id']}",
        headers=_headers(viewer_id),
    )
    assert delete_comment_resp.status_code == 200
    assert delete_comment_resp.json()["status"] == "deleted"

    delete_note_forbidden_resp = client.request(
        "DELETE",
        f"/activities/{note_resp.json()['id']}",
        headers=_headers(claimant_id),
    )
    assert delete_note_forbidden_resp.status_code == 403

    delete_note_admin_resp = client.request(
        "DELETE",
        f"/activities/{note_resp.json()['id']}",
        headers=_headers(1),
    )
    assert delete_note_admin_resp.status_code == 200
    assert delete_note_admin_resp.json()["status"] == "deleted"

    delete_system_event_resp = client.request(
        "DELETE",
        f"/activities/{claim_created_event['id']}",
        headers=_headers(claimant_id),
    )
    assert delete_system_event_resp.status_code == 400

    outsider_task_activity_after_delete_resp = client.get(
        f"/tasks/{task_id}/activities",
        headers=_headers(outsider_id),
    )
    assert outsider_task_activity_after_delete_resp.status_code == 200
    assert outsider_task_activity_after_delete_resp.json() == []

    outsider_claim_activity_resp = client.get(f"/claims/{claim_id}/activities", headers=_headers(outsider_id))
    assert outsider_claim_activity_resp.status_code == 403

    app.dependency_overrides.clear()


def test_task_activity_team_member_requires_claim_id_when_multiple_claims_match(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "TeamActivityReviewer",
            "employee_no": "R971",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    lead_a_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "LeadA", "employee_no": "E973", "department": "RD", "roles": ["employee"]},
    )
    assert lead_a_resp.status_code == 200
    lead_a_id = lead_a_resp.json()["id"]

    lead_b_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "LeadB", "employee_no": "E974", "department": "RD", "roles": ["employee"]},
    )
    assert lead_b_resp.status_code == 200
    lead_b_id = lead_b_resp.json()["id"]

    shared_member_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "SharedMember", "employee_no": "E975", "department": "RD", "roles": ["employee"]},
    )
    assert shared_member_resp.status_code == 200
    shared_member_id = shared_member_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(shared_member_id),
        json={
            "title": "team-activity-ambiguity",
            "scenario": "rd",
            "background": "multiple team claims share one member",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "member must not auto-post to an arbitrary claim",
            "value_reduce_effort": True,
            "value_statement": "force explicit claim selection",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "team-activity-task",
                "goal": "avoid ambiguous claim resolution",
                "scope": "single flow",
                "due_date": (date.today() + timedelta(days=5)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "points": 5,
                "acceptance_criteria": [{"description": "member must provide claim_id", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200
    task_id = review_resp.json()["id"]

    claim_a_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(lead_a_id),
        json={
            "mode": "team",
            "members": [
                {"user_id": lead_a_id, "ratio": 0.5},
                {"user_id": shared_member_id, "ratio": 0.5},
            ],
        },
    )
    assert claim_a_resp.status_code == 200

    claim_b_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(lead_b_id),
        json={
            "mode": "team",
            "members": [
                {"user_id": lead_b_id, "ratio": 0.5},
                {"user_id": shared_member_id, "ratio": 0.5},
            ],
        },
    )
    assert claim_b_resp.status_code == 200
    claim_b_id = claim_b_resp.json()["claim_id"]

    ambiguous_resp = client.post(
        f"/tasks/{task_id}/activities",
        headers=_headers(shared_member_id),
        json={"activity_type": "progress_update", "content": "Working on this now."},
    )
    assert ambiguous_resp.status_code == 400
    assert "claim_id is required" in ambiguous_resp.text

    explicit_resp = client.post(
        f"/tasks/{task_id}/activities",
        headers=_headers(shared_member_id),
        json={
            "activity_type": "progress_update",
            "claim_id": claim_b_id,
            "content": "Posting to the selected team claim.",
        },
    )
    assert explicit_resp.status_code == 200
    assert explicit_resp.json()["claim_id"] == claim_b_id

    claim_b_activity_resp = client.get(f"/claims/{claim_b_id}/activities", headers=_headers(shared_member_id))
    assert claim_b_activity_resp.status_code == 200
    assert any(item["id"] == explicit_resp.json()["id"] for item in claim_b_activity_resp.json())

    app.dependency_overrides.clear()


def test_delete_task_activity_detaches_bound_attachments(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "DetachReviewer",
            "employee_no": "R972",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    author_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "DetachAuthor", "employee_no": "E976", "department": "OPS", "roles": ["employee"]},
    )
    assert author_resp.status_code == 200
    author_id = author_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(author_id),
        json={
            "title": "detach-activity-attachment",
            "scenario": "ops",
            "background": "activity attachments should not be orphaned",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "delete must unbind the file rows",
            "value_reduce_effort": True,
            "value_statement": "preserve attachment accessibility",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "detach-activity-task",
                "goal": "detach activity attachments on delete",
                "scope": "single flow",
                "due_date": (date.today() + timedelta(days=4)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "points": 5,
                "acceptance_criteria": [{"description": "attachment row is detached", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200
    task_id = review_resp.json()["id"]

    upload_resp = client.post(
        "/attachments/upload",
        headers=_headers(author_id),
        files={"file": ("activity.txt", b"activity-attachment", "text/plain")},
    )
    assert upload_resp.status_code == 200
    attachment_id = upload_resp.json()["id"]

    activity_resp = client.post(
        f"/tasks/{task_id}/activities",
        headers=_headers(author_id),
        json={
            "activity_type": "comment",
            "content": "Attached note.",
            "attachment_ids": [attachment_id],
        },
    )
    assert activity_resp.status_code == 200
    activity_id = activity_resp.json()["id"]

    before_delete_attachment_resp = client.get(f"/attachments/{attachment_id}", headers=_headers(author_id))
    assert before_delete_attachment_resp.status_code == 200
    assert before_delete_attachment_resp.json()["entity_type"] == "task_activity"
    assert before_delete_attachment_resp.json()["entity_id"] == activity_id

    delete_resp = client.request("DELETE", f"/activities/{activity_id}", headers=_headers(author_id))
    assert delete_resp.status_code == 200

    after_delete_attachment_resp = client.get(f"/attachments/{attachment_id}", headers=_headers(author_id))
    assert after_delete_attachment_resp.status_code == 200
    assert after_delete_attachment_resp.json()["entity_type"] is None
    assert after_delete_attachment_resp.json()["entity_id"] is None

    download_resp = client.get(f"/attachments/{attachment_id}/download", headers=_headers(author_id))
    assert download_resp.status_code == 200
    assert download_resp.content == b"activity-attachment"

    app.dependency_overrides.clear()


def test_task_milestone_config(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "MilestoneReviewer", "employee_no": "R980", "department": "QA", "roles": ["reviewer", "acceptor", "employee"]},
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    claimant_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "MilestoneClaimant", "employee_no": "E980", "department": "RD", "roles": ["employee"]},
    )
    assert claimant_resp.status_code == 200
    claimant_id = claimant_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(claimant_id),
        json={
            "title": "milestone-config-task",
            "scenario": "rd",
            "background": "need staged delivery",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "validate milestone setup",
            "value_reduce_effort": True,
            "value_statement": "stage work and review",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "complex-milestone-task",
                "goal": "deliver in two stages",
                "scope": "backend only",
                "due_date": (date.today() + timedelta(days=5)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "points": 5,
                "is_complex": True,
                "closing_reward_ratio": 0.4,
                "milestones": [
                    {
                        "sequence": 1,
                        "title": "phase-1",
                        "goal": "ship first increment",
                        "reward_ratio": 0.3,
                        "acceptance_criteria": [{"description": "phase 1 done", "type": "quantified"}],
                    },
                    {
                        "sequence": 2,
                        "title": "phase-2",
                        "goal": "ship second increment",
                        "reward_ratio": 0.3,
                        "acceptance_criteria": [{"description": "phase 2 done", "type": "quantified"}],
                    },
                ],
                "acceptance_criteria": [{"description": "final merged output", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200
    task_id = review_resp.json()["id"]

    milestones_resp = client.get(f"/tasks/{task_id}/milestones", headers=_headers(claimant_id))
    assert milestones_resp.status_code == 200
    milestones_payload = milestones_resp.json()
    assert len(milestones_payload) == 2
    assert [item["sequence"] for item in milestones_payload] == [1, 2]
    assert all(item["status"] == "pending" for item in milestones_payload)

    claim_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(claimant_id),
        json={"mode": "individual"},
    )
    assert claim_resp.status_code == 200

    milestones_after_claim_resp = client.get(f"/tasks/{task_id}/milestones", headers=_headers(claimant_id))
    assert milestones_after_claim_resp.status_code == 200
    active_like = [
        item
        for item in milestones_after_claim_resp.json()
        if item["status"] in {"active", "pending_acceptance"}
    ]
    assert len(active_like) == 1
    assert active_like[0]["sequence"] == 1

    simple_problem_resp = client.post(
        "/problems",
        headers=_headers(claimant_id),
        json={
            "title": "simple-task-should-reject-milestones",
            "scenario": "rd",
            "background": "simple flow",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "simple tasks cannot carry milestones",
            "value_reduce_effort": True,
            "value_statement": "validate guard",
        },
    )
    assert simple_problem_resp.status_code == 200
    simple_problem_id = simple_problem_resp.json()["id"]

    simple_review_resp = client.post(
        f"/problems/{simple_problem_id}/review",
        headers=_headers(reviewer_id),
        json={
            "approve": True,
            "task": {
                "title": "simple-task",
                "goal": "no milestones",
                "scope": "single delivery",
                "due_date": (date.today() + timedelta(days=5)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "points": 5,
                "is_complex": False,
                "milestones": [
                    {
                        "sequence": 1,
                        "title": "invalid-phase",
                        "goal": "should fail",
                        "reward_ratio": 0.5,
                        "acceptance_criteria": [{"description": "invalid", "type": "quantified"}],
                    }
                ],
                "acceptance_criteria": [{"description": "done", "type": "quantified"}],
            },
        },
    )
    assert simple_review_resp.status_code == 422

    ratio_problem_resp = client.post(
        "/problems",
        headers=_headers(claimant_id),
        json={
            "title": "milestone-ratio-guard",
            "scenario": "rd",
            "background": "ratio validation",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "ratios must add up to one with closing",
            "value_reduce_effort": True,
            "value_statement": "validate ratio sum guard",
        },
    )
    assert ratio_problem_resp.status_code == 200
    ratio_problem_id = ratio_problem_resp.json()["id"]

    bad_ratio_review_resp = client.post(
        f"/problems/{ratio_problem_id}/review",
        headers=_headers(reviewer_id),
        json={
            "approve": True,
            "task": {
                "title": "bad-ratio-task",
                "goal": "invalid ratios",
                "scope": "single delivery",
                "due_date": (date.today() + timedelta(days=5)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "points": 5,
                "is_complex": True,
                "closing_reward_ratio": 0.3,
                "milestones": [
                    {
                        "sequence": 1,
                        "title": "phase-1",
                        "goal": "phase 1",
                        "reward_ratio": 0.4,
                        "acceptance_criteria": [{"description": "phase 1", "type": "quantified"}],
                    },
                    {
                        "sequence": 2,
                        "title": "phase-2",
                        "goal": "phase 2",
                        "reward_ratio": 0.4,
                        "acceptance_criteria": [{"description": "phase 2", "type": "quantified"}],
                    },
                ],
                "acceptance_criteria": [{"description": "done", "type": "quantified"}],
            },
        },
    )
    assert bad_ratio_review_resp.status_code == 422

    app.dependency_overrides.clear()


def test_milestone_execution(tmp_path: Path) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "MilestoneExecReviewer", "employee_no": "R981", "department": "QA", "roles": ["reviewer", "acceptor", "employee"]},
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    claimant_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "MilestoneExecClaimant", "employee_no": "E981", "department": "RD", "roles": ["employee"]},
    )
    assert claimant_resp.status_code == 200
    claimant_id = claimant_resp.json()["id"]

    problem_resp = client.post(
        "/problems",
        headers=_headers(claimant_id),
        json={
            "title": "milestone-execution-task",
            "scenario": "rd",
            "background": "execute milestones",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "cover submit and accept lifecycle",
            "value_reduce_effort": True,
            "value_statement": "staged delivery validation",
        },
    )
    assert problem_resp.status_code == 200
    problem_id = problem_resp.json()["id"]

    review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        problem_id,
        {
            "approve": True,
            "task": {
                "title": "milestone-execution",
                "goal": "complete staged delivery",
                "scope": "milestone submit/accept then final deliverable",
                "due_date": (date.today() + timedelta(days=5)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "points": 5,
                "is_complex": True,
                "closing_reward_ratio": 0.4,
                "milestones": [
                    {
                        "sequence": 1,
                        "title": "phase-1",
                        "goal": "first checkpoint",
                        "reward_ratio": 0.3,
                        "acceptance_criteria": [{"description": "phase 1 done", "type": "quantified"}],
                    },
                    {
                        "sequence": 2,
                        "title": "phase-2",
                        "goal": "second checkpoint",
                        "reward_ratio": 0.3,
                        "acceptance_criteria": [{"description": "phase 2 done", "type": "quantified"}],
                    },
                ],
                "acceptance_criteria": [{"description": "final delivery done", "type": "quantified"}],
            },
        },
    )
    assert review_resp.status_code == 200
    task_id = review_resp.json()["id"]

    claim_resp = client.post(
        f"/tasks/{task_id}/claims",
        headers=_headers(claimant_id),
        json={"mode": "individual"},
    )
    assert claim_resp.status_code == 200
    claim_id = claim_resp.json()["claim_id"]

    milestones_resp = client.get(f"/tasks/{task_id}/milestones", headers=_headers(claimant_id))
    assert milestones_resp.status_code == 200
    milestones = milestones_resp.json()
    first_id = milestones[0]["id"]
    second_id = milestones[1]["id"]
    assert milestones[0]["status"] == "active"

    submit_first_resp = client.post(
        f"/milestones/{first_id}/submit",
        headers=_headers(claimant_id),
        json={
            "claim_id": claim_id,
            "summary": "phase-1 output",
            "criteria_results": ["phase 1 met"],
            "evidence_urls": [],
        },
    )
    assert submit_first_resp.status_code == 200
    assert submit_first_resp.json()["status"] == "pending_acceptance"

    pending_first_resp = client.get("/milestones/pending-acceptance/mine", headers=_headers(reviewer_id))
    assert pending_first_resp.status_code == 200
    assert any(item["milestone_id"] == first_id for item in pending_first_resp.json())

    rework_resp = client.post(
        f"/milestones/{first_id}/accept",
        headers=_headers(reviewer_id),
        json={"result": "rework", "comment": "please refine"},
    )
    assert rework_resp.status_code == 200
    assert rework_resp.json()["status"] == "rework"

    submit_first_again_resp = client.post(
        f"/milestones/{first_id}/submit",
        headers=_headers(claimant_id),
        json={
            "claim_id": claim_id,
            "summary": "phase-1 output v2",
            "criteria_results": ["phase 1 met again"],
            "evidence_urls": [],
        },
    )
    assert submit_first_again_resp.status_code == 200
    assert submit_first_again_resp.json()["status"] == "pending_acceptance"

    approve_first_resp = client.post(
        f"/milestones/{first_id}/accept",
        headers=_headers(reviewer_id),
        json={"result": "approved", "comment": "phase 1 approved"},
    )
    assert approve_first_resp.status_code == 200
    assert approve_first_resp.json()["status"] == "approved"
    assert approve_first_resp.json()["next_milestone_id"] == second_id

    final_before_second_resp = client.post(
        f"/claims/{claim_id}/deliverables",
        headers=_headers(claimant_id),
        json={"summary": "too early", "criteria_results": ["not ready"], "evidence_urls": []},
    )
    assert final_before_second_resp.status_code == 400

    submit_second_resp = client.post(
        f"/milestones/{second_id}/submit",
        headers=_headers(claimant_id),
        json={
            "claim_id": claim_id,
            "summary": "phase-2 output",
            "criteria_results": ["phase 2 met"],
            "evidence_urls": [],
        },
    )
    assert submit_second_resp.status_code == 200
    assert submit_second_resp.json()["status"] == "pending_acceptance"

    approve_second_resp = client.post(
        f"/milestones/{second_id}/accept",
        headers=_headers(reviewer_id),
        json={"result": "approved", "comment": "phase 2 approved"},
    )
    assert approve_second_resp.status_code == 200
    assert approve_second_resp.json()["status"] == "approved"
    assert approve_second_resp.json()["next_milestone_id"] is None

    final_submit_resp = client.post(
        f"/claims/{claim_id}/deliverables",
        headers=_headers(claimant_id),
        json={"summary": "all milestones done", "criteria_results": ["final ready"], "evidence_urls": []},
    )
    assert final_submit_resp.status_code == 200
    deliverable_id = final_submit_resp.json()["deliverable_id"]

    final_accept_resp = client.post(
        f"/deliverables/{deliverable_id}/accept",
        headers=_headers(reviewer_id),
        json={"result": "approved", "comment": "final approved"},
    )
    assert final_accept_resp.status_code == 200
    assert final_accept_resp.json()["task_status"] == "completed"

    self_acceptor_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "SelfAcceptor", "employee_no": "E982", "department": "RD", "roles": ["employee", "acceptor"]},
    )
    assert self_acceptor_resp.status_code == 200
    self_acceptor_id = self_acceptor_resp.json()["id"]

    self_problem_resp = client.post(
        "/problems",
        headers=_headers(self_acceptor_id),
        json={
            "title": "self-accept-guard",
            "scenario": "rd",
            "background": "self accepter should be blocked",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "accepter cannot accept own milestone output",
            "value_reduce_effort": True,
            "value_statement": "guard self acceptance",
        },
    )
    assert self_problem_resp.status_code == 200
    self_problem_id = self_problem_resp.json()["id"]

    self_task_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        self_problem_id,
        {
            "approve": True,
            "task": {
                "title": "self-accept-task",
                "goal": "guard self accept",
                "scope": "single flow",
                "due_date": (date.today() + timedelta(days=5)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": self_acceptor_id,
                "points": 5,
                "is_complex": True,
                "closing_reward_ratio": 0.4,
                "milestones": [
                    {
                        "sequence": 1,
                        "title": "phase-1",
                        "goal": "first",
                        "reward_ratio": 0.3,
                        "acceptance_criteria": [{"description": "phase 1", "type": "quantified"}],
                    },
                    {
                        "sequence": 2,
                        "title": "phase-2",
                        "goal": "second",
                        "reward_ratio": 0.3,
                        "acceptance_criteria": [{"description": "phase 2", "type": "quantified"}],
                    },
                ],
                "acceptance_criteria": [{"description": "final", "type": "quantified"}],
            },
        },
    )
    assert self_task_resp.status_code == 200
    self_task_id = self_task_resp.json()["id"]

    self_claim_resp = client.post(
        f"/tasks/{self_task_id}/claims",
        headers=_headers(self_acceptor_id),
        json={"mode": "individual"},
    )
    assert self_claim_resp.status_code == 200
    self_claim_id = self_claim_resp.json()["claim_id"]

    self_milestones_resp = client.get(f"/tasks/{self_task_id}/milestones", headers=_headers(self_acceptor_id))
    assert self_milestones_resp.status_code == 200
    self_first_milestone_id = self_milestones_resp.json()[0]["id"]

    self_submit_resp = client.post(
        f"/milestones/{self_first_milestone_id}/submit",
        headers=_headers(self_acceptor_id),
        json={
            "claim_id": self_claim_id,
            "summary": "self output",
            "criteria_results": ["done"],
            "evidence_urls": [],
        },
    )
    assert self_submit_resp.status_code == 200

    self_accept_resp = client.post(
        f"/milestones/{self_first_milestone_id}/accept",
        headers=_headers(self_acceptor_id),
        json={"result": "approved", "comment": "self should fail"},
    )
    assert self_accept_resp.status_code == 403

    app.dependency_overrides.clear()


def test_stale_progress_reminders_cover_simple_and_complex_claims(
    tmp_path: Path,
    monkeypatch,
) -> None:
    client = _setup_client(tmp_path)

    reviewer_resp = client.post(
        "/users",
        headers=_headers(1),
        json={
            "name": "StaleReminderReviewer",
            "employee_no": "R983",
            "department": "QA",
            "roles": ["reviewer", "acceptor", "employee"],
        },
    )
    assert reviewer_resp.status_code == 200
    reviewer_id = reviewer_resp.json()["id"]

    simple_user_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "SimpleOwner", "employee_no": "E983", "department": "RD", "roles": ["employee"]},
    )
    assert simple_user_resp.status_code == 200
    simple_user_id = simple_user_resp.json()["id"]

    complex_user_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "ComplexOwner", "employee_no": "E984", "department": "RD", "roles": ["employee"]},
    )
    assert complex_user_resp.status_code == 200
    complex_user_id = complex_user_resp.json()["id"]

    fresh_user_resp = client.post(
        "/users",
        headers=_headers(1),
        json={"name": "FreshOwner", "employee_no": "E985", "department": "RD", "roles": ["employee"]},
    )
    assert fresh_user_resp.status_code == 200
    fresh_user_id = fresh_user_resp.json()["id"]

    simple_problem_resp = client.post(
        "/problems",
        headers=_headers(simple_user_id),
        json={
            "title": "stale-simple-problem",
            "scenario": "rd",
            "background": "simple task reminder",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "simple claim should be reminded when no progress",
            "value_reduce_effort": True,
            "value_statement": "cover simple stale reminder",
        },
    )
    assert simple_problem_resp.status_code == 200
    simple_problem_id = simple_problem_resp.json()["id"]

    simple_review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        simple_problem_id,
        {
            "approve": True,
            "task": {
                "title": "stale-simple-task",
                "goal": "simple stale reminder",
                "scope": "single flow",
                "due_date": (date.today() + timedelta(days=7)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "points": 5,
                "acceptance_criteria": [{"description": "simple stale reminder works", "type": "quantified"}],
            },
        },
    )
    assert simple_review_resp.status_code == 200
    simple_task_id = simple_review_resp.json()["id"]

    complex_problem_resp = client.post(
        "/problems",
        headers=_headers(complex_user_id),
        json={
            "title": "stale-complex-problem",
            "scenario": "rd",
            "background": "complex task reminder",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "complex claim should be reminded when no progress",
            "value_reduce_effort": True,
            "value_statement": "cover complex stale reminder",
        },
    )
    assert complex_problem_resp.status_code == 200
    complex_problem_id = complex_problem_resp.json()["id"]

    complex_review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        complex_problem_id,
        {
            "approve": True,
            "task": {
                "title": "stale-complex-task",
                "goal": "complex stale reminder",
                "scope": "single flow",
                "due_date": (date.today() + timedelta(days=7)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "points": 5,
                "is_complex": True,
                "closing_reward_ratio": 0.4,
                "milestones": [
                    {
                        "sequence": 1,
                        "title": "phase-1",
                        "goal": "first phase",
                        "reward_ratio": 0.3,
                        "acceptance_criteria": [{"description": "phase 1 done", "type": "quantified"}],
                    },
                    {
                        "sequence": 2,
                        "title": "phase-2",
                        "goal": "second phase",
                        "reward_ratio": 0.3,
                        "acceptance_criteria": [{"description": "phase 2 done", "type": "quantified"}],
                    },
                ],
                "acceptance_criteria": [{"description": "complex stale reminder works", "type": "quantified"}],
            },
        },
    )
    assert complex_review_resp.status_code == 200
    complex_task_id = complex_review_resp.json()["id"]

    fresh_problem_resp = client.post(
        "/problems",
        headers=_headers(fresh_user_id),
        json={
            "title": "fresh-progress-problem",
            "scenario": "rd",
            "background": "recent progress should suppress reminder",
            "frequency": "weekly",
            "impact_scope": "team",
            "description": "fresh claim should not be reminded",
            "value_reduce_effort": True,
            "value_statement": "avoid false reminder",
        },
    )
    assert fresh_problem_resp.status_code == 200
    fresh_problem_id = fresh_problem_resp.json()["id"]

    fresh_review_resp = _review_problem_and_create_task(
        client,
        reviewer_id,
        fresh_problem_id,
        {
            "approve": True,
            "task": {
                "title": "fresh-progress-task",
                "goal": "fresh progress exclusion",
                "scope": "single flow",
                "due_date": (date.today() + timedelta(days=7)).isoformat(),
                "level": "C",
                "reward_total": 300,
                "proposer_ratio": 0.2,
                "accepter_id": reviewer_id,
                "points": 5,
                "acceptance_criteria": [{"description": "fresh progress is ignored", "type": "quantified"}],
            },
        },
    )
    assert fresh_review_resp.status_code == 200
    fresh_task_id = fresh_review_resp.json()["id"]

    simple_claim_resp = client.post(
        f"/tasks/{simple_task_id}/claims",
        headers=_headers(simple_user_id),
        json={"mode": "individual"},
    )
    assert simple_claim_resp.status_code == 200
    simple_claim_id = simple_claim_resp.json()["claim_id"]

    complex_claim_resp = client.post(
        f"/tasks/{complex_task_id}/claims",
        headers=_headers(complex_user_id),
        json={"mode": "individual"},
    )
    assert complex_claim_resp.status_code == 200
    complex_claim_id = complex_claim_resp.json()["claim_id"]

    fresh_claim_resp = client.post(
        f"/tasks/{fresh_task_id}/claims",
        headers=_headers(fresh_user_id),
        json={"mode": "individual"},
    )
    assert fresh_claim_resp.status_code == 200
    fresh_claim_id = fresh_claim_resp.json()["claim_id"]

    fresh_progress_resp = client.post(
        f"/tasks/{fresh_task_id}/activities",
        headers=_headers(fresh_user_id),
        json={
            "activity_type": "progress_update",
            "claim_id": fresh_claim_id,
            "content": "Recent progress update.",
        },
    )
    assert fresh_progress_resp.status_code == 200

    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", connect_args={"check_same_thread": False})
    stale_created_at = datetime.utcnow() - timedelta(days=5)
    with Session(engine) as session:
        simple_claim = session.get(Claim, simple_claim_id)
        complex_claim = session.get(Claim, complex_claim_id)
        assert simple_claim is not None
        assert complex_claim is not None
        simple_claim.created_at = stale_created_at
        complex_claim.created_at = stale_created_at
        session.commit()

    notification_calls: list[dict] = []

    def _fake_notify(session, **kwargs):
        notification_calls.append(kwargs)
        return {"status": "mocked"}

    monkeypatch.setattr("app.services_task_activity.notify_stale_progress_reminder", _fake_notify)

    with Session(engine) as session:
        first_job = run_stale_progress_reminders(session, actor_id=reviewer_id, now=datetime.utcnow())
    assert first_job["reminders_created"] == 2
    assert set(first_job["notified_claim_ids"]) == {simple_claim_id, complex_claim_id}
    assert len(notification_calls) == 2
    assert {call["claim_id"] for call in notification_calls} == {simple_claim_id, complex_claim_id}

    with Session(engine) as session:
        second_job = run_stale_progress_reminders(session, actor_id=reviewer_id, now=datetime.utcnow())
    assert second_job["reminders_created"] == 0

    simple_activity_resp = client.get(f"/claims/{simple_claim_id}/activities", headers=_headers(simple_user_id))
    assert simple_activity_resp.status_code == 200
    assert any(
        item["activity_type"] == "system_event"
        and item["detail"].get("event_key") == "stale_progress_reminder"
        for item in simple_activity_resp.json()
    )

    complex_activity_resp = client.get(f"/claims/{complex_claim_id}/activities", headers=_headers(complex_user_id))
    assert complex_activity_resp.status_code == 200
    assert any(
        item["activity_type"] == "system_event"
        and item["detail"].get("event_key") == "stale_progress_reminder"
        for item in complex_activity_resp.json()
    )

    fresh_activity_resp = client.get(f"/claims/{fresh_claim_id}/activities", headers=_headers(fresh_user_id))
    assert fresh_activity_resp.status_code == 200
    assert all(
        item["detail"].get("event_key") != "stale_progress_reminder"
        for item in fresh_activity_resp.json()
        if item["activity_type"] == "system_event"
    )

    app.dependency_overrides.clear()
